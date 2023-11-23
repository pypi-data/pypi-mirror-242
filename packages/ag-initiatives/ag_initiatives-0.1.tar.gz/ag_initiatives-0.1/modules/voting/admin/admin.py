import copy
import io
import zipfile
from collections import defaultdict
from typing import List

import nested_admin
import openpyxl
from django import forms
from django.contrib import admin
from django.forms import TextInput
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

from config.settings import BASE_DIR
from modules.core.models import Category, User
from modules.voting.models import UserVote, ImportXlsModel, RejectReason
from modules.voting.models import VoteQuestion, VoteAnswerOption, Vote


# TODO Разнести содержимое файла по отдельным файлам, как в core

class VoteAnswerOptionAdminForm(forms.ModelForm):
    class Meta:
        model = VoteAnswerOption
        fields = '__all__'
        widgets = {
            'brief': TextInput(attrs={'size': 70})
        }


class VoteQuestionAdminForm(forms.ModelForm):
    class Meta:
        model = VoteQuestion
        fields = '__all__'
        widgets = {
            'brief': TextInput(attrs={'size': 70}),
            'name': TextInput(attrs={'size': 70})
        }


class VoteAdminForm(forms.ModelForm):
    class Meta:
        model = Vote
        fields = '__all__'
        widgets = {
            'topic': TextInput(attrs={'size': 70}),
            'name': TextInput(attrs={'size': 70})
        }


class CategoryForm(forms.ModelForm):
    class Meta:
        model = Category
        fields = '__all__'
        widgets = {
            'name': TextInput(attrs={'size': 70})
        }


class VoteAnswerOptionAdmin(nested_admin.NestedStackedInline):
    form = VoteAnswerOptionAdminForm
    model = VoteAnswerOption
    extra = 0


class VoteQuestionAdmin(nested_admin.NestedStackedInline):
    form = VoteQuestionAdminForm
    model = VoteQuestion
    extra = 0
    inlines = (VoteAnswerOptionAdmin,)


# TODO refactor and move this somewhere
def get_vote_data_for_report(vote):
    user_votes = UserVote.objects.filter(vote=vote).order_by('locality', 'question', 'answer_option').select_related(
        'user',
    )
    localiies_map = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int)))))

    year_now = timezone.now().date().year
    for user_vote in user_votes:
        if not user_vote.answer_option:
            continue
        locality_id = user_vote.locality.pk
        questions_id = user_vote.question.pk
        answer_id = user_vote.answer_option.pk

        user_age_group = 0
        if user_vote.user.birth_date:
            user_age = year_now - user_vote.user.birth_date.year
            if user_age < 18:
                user_age_group = 1
            elif user_age < 25:
                user_age_group = 2
            elif user_age < 35:
                user_age_group = 3
            elif user_age < 45:
                user_age_group = 4
            elif user_age < 55:
                user_age_group = 5
            elif user_age >= 55:
                user_age_group = 6

        localiies_map[locality_id][questions_id][answer_id][user_vote.user.gender][user_age_group] += 1
        localiies_map['all_answers'][questions_id][answer_id][user_vote.user.gender][user_age_group] += 1

    return localiies_map


# TODO refactor and move this somewhere
def create_vote_report_xlsx(vote, vote_results_data):
    wb = Workbook()

    ws = wb.active

    ws.column_dimensions['A'].width = 48

    def question_total_by_age_category(data, age_category):
        result = 0
        for answer in data.values():
            result += answer['M'][age_category]
            result += answer['F'][age_category]
        return result

    def get_vote_str_localities(obj: Vote) -> str:
        localities = obj.locality.all()
        localities_str = ""
        for locality in localities:
            localities_str += f"{str(locality)}, "
        if localities_str != "":
            localities_str = localities_str[:-2]
        return localities_str

    try:
        vote_organization = str(vote.department.name)
    except AttributeError:
        vote_organization = "Не указан"
    vote_name = str(vote.name)
    vote_date = f"{vote.start_date} - {vote.end_date}"
    vote_locality = get_vote_str_localities(vote)
    vote_count = vote.uservote_set.filter(locality__in=vote.locality.all()).distinct("user").count()

    vote_date = {
        1: {"name": "Иницииатор голосования:", "value": vote_organization},
        2: {"name": "Наименование голосования:", "value": vote_name},
        3: {"name": "Даты проведения:", "value": vote_date},
        4: {"name": "Территория:", "value": vote_locality},
        5: {"name": "Количество проголосовавших:", "value": vote_count},
    }
    rn = 1
    for i in range(1, 6):
        ws.cell(i, 1, value=vote_date[i]["name"])
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
        ws.cell(i, 2, value=vote_date[i]["value"])
        rn += 1

    rn = 8
    if vote.locality.count() > 1:
        ws.cell(row=rn, column=2, value='Все районы')
        rn += 1
        for question_num, question in enumerate(vote.questions.all()):
            ws.cell(row=rn, column=2, value='Вопрос № {}; {}'.format(question.id, question.brief))
            rn += 1
            ws.cell(row=rn, column=1, value='Вариант ответа')
            ws.cell(row=rn, column=2, value='Пол')
            ws.cell(row=rn, column=3, value='до 18')
            ws.cell(row=rn, column=4, value='18-24')
            ws.cell(row=rn, column=5, value='25-34')
            ws.cell(row=rn, column=6, value='35-44')
            ws.cell(row=rn, column=7, value='45 -54')
            ws.cell(row=rn, column=8, value='55 и старше')
            ws.cell(row=rn, column=9, value='Итог')
            rn += 1
            for answer_num, answer in enumerate(question.answers.all()):
                ws.cell(row=rn, column=1, value='{}'.format(answer.brief))
                ws.cell(row=rn + 0, column=2, value='М')
                ws.cell(row=rn + 1, column=2, value='Ж')

                ws.cell(row=rn + 0, column=3,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['M'][1]))
                ws.cell(row=rn + 1, column=3,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['F'][1]))

                ws.cell(row=rn + 0, column=4,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['M'][2]))
                ws.cell(row=rn + 1, column=4,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['F'][2]))

                ws.cell(row=rn + 0, column=5,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['M'][3]))
                ws.cell(row=rn + 1, column=5,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['F'][3]))

                ws.cell(row=rn + 0, column=6,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['M'][4]))
                ws.cell(row=rn + 1, column=6,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['F'][4]))

                ws.cell(row=rn + 0, column=7,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['M'][5]))
                ws.cell(row=rn + 1, column=7,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['F'][5]))

                ws.cell(row=rn + 0, column=8,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['M'][6]))
                ws.cell(row=rn + 1, column=8,
                        value='{}'.format(vote_results_data['all_answers'][question.pk][answer.pk]['F'][6]))

                m_total = sum(vote_results_data['all_answers'][question.pk][answer.pk]['M'].values())
                f_total = sum(vote_results_data['all_answers'][question.pk][answer.pk]['F'].values())
                ws.cell(row=rn + 0, column=9, value='{}'.format(m_total))
                ws.cell(row=rn + 1, column=9, value='{}'.format(f_total))
                rn += 2
            ws.cell(row=rn, column=1, value='Итог')
            ws.cell(row=rn, column=3,
                    value='{}'.format(question_total_by_age_category(vote_results_data['all_answers'][question.pk], 1)))
            ws.cell(row=rn, column=4,
                    value='{}'.format(question_total_by_age_category(vote_results_data['all_answers'][question.pk], 2)))
            ws.cell(row=rn, column=5,
                    value='{}'.format(question_total_by_age_category(vote_results_data['all_answers'][question.pk], 3)))
            ws.cell(row=rn, column=6,
                    value='{}'.format(question_total_by_age_category(vote_results_data['all_answers'][question.pk], 4)))
            ws.cell(row=rn, column=7,
                    value='{}'.format(question_total_by_age_category(vote_results_data['all_answers'][question.pk], 5)))
            ws.cell(row=rn, column=8,
                    value='{}'.format(question_total_by_age_category(vote_results_data['all_answers'][question.pk], 6)))
            rn += 2
        rn += 2

    for locality_num, locality in enumerate(vote.locality.all()):
        ws.cell(row=rn, column=2, value='Опрос от ({}-{}) по {}'.format(vote.start_date, vote.end_date, locality.name))
        rn += 1
        for question_num, question in enumerate(vote.questions.all()):
            ws.cell(row=rn, column=2, value='Вопрос № {}; {}'.format(question.id, question.brief))
            rn += 1
            ws.cell(row=rn, column=1, value='Вариант ответа')
            ws.cell(row=rn, column=2, value='Пол')
            ws.cell(row=rn, column=3, value='до 18')
            ws.cell(row=rn, column=4, value='18-24')
            ws.cell(row=rn, column=5, value='25-34')
            ws.cell(row=rn, column=6, value='35-44')
            ws.cell(row=rn, column=7, value='45 -54')
            ws.cell(row=rn, column=8, value='55 и старше')
            ws.cell(row=rn, column=9, value='Итог')
            rn += 1
            for answer_num, answer in enumerate(question.answers.all()):
                ws.cell(row=rn, column=1, value='{}'.format(answer.brief))
                ws.cell(row=rn + 0, column=2, value='М')
                ws.cell(row=rn + 1, column=2, value='Ж')

                ws.cell(row=rn + 0, column=3,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['M'][1]))
                ws.cell(row=rn + 1, column=3,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['F'][1]))

                ws.cell(row=rn + 0, column=4,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['M'][2]))
                ws.cell(row=rn + 1, column=4,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['F'][2]))

                ws.cell(row=rn + 0, column=5,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['M'][3]))
                ws.cell(row=rn + 1, column=5,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['F'][3]))

                ws.cell(row=rn + 0, column=6,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['M'][4]))
                ws.cell(row=rn + 1, column=6,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['F'][4]))

                ws.cell(row=rn + 0, column=7,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['M'][5]))
                ws.cell(row=rn + 1, column=7,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['F'][5]))

                ws.cell(row=rn + 0, column=8,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['M'][6]))
                ws.cell(row=rn + 1, column=8,
                        value='{}'.format(vote_results_data[locality.pk][question.pk][answer.pk]['F'][6]))

                m_total = sum(vote_results_data[locality.pk][question.pk][answer.pk]['M'].values())
                f_total = sum(vote_results_data[locality.pk][question.pk][answer.pk]['F'].values())
                ws.cell(row=rn + 0, column=9, value='{}'.format(m_total))
                ws.cell(row=rn + 1, column=9, value='{}'.format(f_total))
                rn += 2
            ws.cell(row=rn, column=1, value='Итог')
            ws.cell(row=rn, column=3,
                    value='{}'.format(question_total_by_age_category(vote_results_data[locality.pk][question.pk], 1)))
            ws.cell(row=rn, column=4,
                    value='{}'.format(question_total_by_age_category(vote_results_data[locality.pk][question.pk], 2)))
            ws.cell(row=rn, column=5,
                    value='{}'.format(question_total_by_age_category(vote_results_data[locality.pk][question.pk], 3)))
            ws.cell(row=rn, column=6,
                    value='{}'.format(question_total_by_age_category(vote_results_data[locality.pk][question.pk], 4)))
            ws.cell(row=rn, column=7,
                    value='{}'.format(question_total_by_age_category(vote_results_data[locality.pk][question.pk], 5)))
            ws.cell(row=rn, column=8,
                    value='{}'.format(question_total_by_age_category(vote_results_data[locality.pk][question.pk], 6)))
            rn += 2
        rn += 2

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    return xlsx_buffer


# TODO refactor and move this somewhere
def vote_report(modeladmin, request, queryset):
    if queryset.count() == 0:
        return

    xlsx_files = []
    for vote in queryset:
        vote_results_data = get_vote_data_for_report(vote)
        xlsx_buffer = create_vote_report_xlsx(vote, vote_results_data)
        filename = 'vote{}.xls'.format(vote.id)
        xlsx_files.append((filename, xlsx_buffer))

    response = None
    if len(xlsx_files) == 1:
        filename, xlsx_buffer = xlsx_files[0]
        response = HttpResponse(content_type='application/excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        response.write(xlsx_buffer.getvalue())
    else:
        zip_buffer = io.BytesIO()
        zip_archive = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)
        for filename, xlsx_buffer in xlsx_files:
            zip_archive.writestr(filename, xlsx_buffer.getvalue())
        zip_archive.close()
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format('vote_results.zip')

    return response


def create_vote_report2_xlsx(vote: Vote):
    def get_vote_str_localities(obj: Vote) -> str:
        localities = obj.locality.all()
        if len(localities) == 0:
            return 'aaaaaa'
        if len(localities) == 1:
            localities_str = str(localities.first())
        else:
            localities_str = ''
            for locality in localities:
                localities_str += f'{str(locality)}, '

            localities_str = localities_str[:-2]

        return localities_str

    wb = openpyxl.load_workbook(f'{BASE_DIR}/modules/voting/admin/report2.xlsx')
    ws = wb.worksheets[0]

    table_cell = ws.cell(9, 2)
    table_font = copy.copy(table_cell.font)
    table_border = copy.copy(table_cell.border)
    table_alignment = copy.copy(table_cell.alignment)

    user_votes = UserVote.objects.filter(vote=vote)

    vote_organization = str(vote.department.name)
    vote_name = str(vote.name)
    vote_date = f'{vote.start_date} - {vote.end_date}'
    vote_locality = get_vote_str_localities(vote)
    vote_count = vote.uservote_set.filter(locality__in=vote.locality.all()).distinct('user').count()
    ws.cell(2, 3, value=vote_organization)
    ws.cell(3, 3, value=vote_name)
    ws.cell(4, 3, value=vote_date)
    ws.cell(5, 3, value=vote_locality)
    ws.cell(6, 3, value=vote_count)

    localities = vote.locality.all()

    current_row = 11
    last_column = 0

    locality_table_row_number = 1
    for locality in localities:
        locality_user_votes = vote.uservote_set.filter(locality=locality).distinct('user')
        locality_votes = user_votes.filter(locality=locality)
        ws.cell(current_row, 1, value=locality_table_row_number)
        ws.cell(current_row, 2, value=str(locality))
        user_age_counts = {
            1: 0,
            2: 0,
            3: 0,
            4: 0,
            5: 0,
            6: 0,
        }
        current_column = 9
        for locality_vote in locality_user_votes:
            if locality_vote.user.birth_date:
                user_age = locality_vote.user.age
                if user_age < 18:
                    user_age_group = 1
                elif user_age < 30:
                    user_age_group = 2
                elif user_age < 50:
                    user_age_group = 3
                elif user_age < 65:
                    user_age_group = 4
                elif user_age >= 65:
                    user_age_group = 5
                else:
                    continue
                user_age_counts[user_age_group] += 1

        ws.cell(current_row, 3, value=locality_user_votes.count())
        ws.cell(current_row, 4, value=user_age_counts[1])
        ws.cell(current_row, 5, value=user_age_counts[2])
        ws.cell(current_row, 6, value=user_age_counts[3])
        ws.cell(current_row, 7, value=user_age_counts[4])
        ws.cell(current_row, 8, value=user_age_counts[5])

        # заполнение статистики по вариантам ответа
        question_index = 1
        for question in vote.questions.order_by('id').all():
            try:
                ws.cell(9, current_column, value=f'Вопрос {question_index}: {question.brief}')
            except AttributeError:
                pass
            question_index += 1

            ws.cell(10, current_column, value='Всего')
            ws.cell(current_row, current_column, value=locality_votes.filter(question=question).count())
            current_column += 1
            for answer_option in question.answers.order_by('id').all():
                ws.cell(10, current_column, value=answer_option.brief)

                user_vote_for_answer_option_count = locality_user_votes.filter(answer_option=answer_option).count()

                ws.cell(current_row, current_column, value=user_vote_for_answer_option_count)
                current_column += 1

        current_row += 1
        locality_table_row_number += 1
        last_column = current_column

    # применение стилей ко всем ячейкам
    for row in range(9, current_row):
        for column in range(1, last_column):
            cell = ws.cell(row, column)
            cell.font = table_font
            cell.border = table_border
            cell.alignment = table_alignment

    for column_index in range(9, last_column):
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = 19

    # Объединение ячеек для вопросов
    start_column_index_to_merge = 9
    for column_index in range(start_column_index_to_merge, last_column):

        if ws.cell(9, column_index + 1).value:
            ws.merge_cells(start_row=9, end_row=9, start_column=start_column_index_to_merge, end_column=column_index)
            start_column_index_to_merge = column_index + 1
    ws.merge_cells(start_row=9, end_row=9, start_column=start_column_index_to_merge, end_column=last_column - 1)

    ws2 = wb.worksheets[1]

    ws2.cell(1, 1, value=f'Сводный список ответов из поля «другое». Информация на: {timezone.now().date()}')

    custom_user_votes = user_votes.filter(custom_answer__isnull=False) \
        .order_by('locality__id').select_related('question')

    row_index = 5
    for user_vote in custom_user_votes:
        cells_to_format: List[Cell] = [
            ws2.cell(row_index, 1, value=str(user_vote.locality)),
            ws2.cell(row_index, 2, value=user_vote.question.brief),
            ws2.cell(row_index, 3, value=user_vote.custom_answer)
        ]

        for cell_to_format in cells_to_format:
            cell_to_format.font = table_font
            cell_to_format.border = table_border
            cell_to_format.alignment = table_alignment

        row_index += 1

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    return xlsx_buffer


def vote_report2(modeladmin, request, queryset):
    if queryset.count() == 0:
        return

    xlsx_files = []
    for vote in queryset:
        xlsx_buffer = create_vote_report2_xlsx(vote)
        filename = 'vote{}.xls'.format(vote.id)
        xlsx_files.append((filename, xlsx_buffer))

    response = None
    if len(xlsx_files) == 1:
        filename, xlsx_buffer = xlsx_files[0]
        response = HttpResponse(content_type='application/excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        response.write(xlsx_buffer.getvalue())
    else:
        zip_buffer = io.BytesIO()
        zip_archive = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)
        for filename, xlsx_buffer in xlsx_files:
            zip_archive.writestr(filename, xlsx_buffer.getvalue())
        zip_archive.close()
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format('vote_results.zip')

    return response


vote_report.short_description = 'Отчет по голосованию'
vote_report2.short_description = 'Отчет по голосованию 2'


@admin.register(Vote)
class VoteAdmin(nested_admin.NestedModelAdmin):
    form = VoteAdminForm
    list_display = [
        'name',
        'department',
        'category',
        'is_opened',
        'is_published',
        'start_date',
        'end_date',
        'voted_users_count',
    ]
    inlines = (VoteQuestionAdmin,)
    actions = [vote_report, vote_report2]

    def voted_users_count(self, instance):
        return instance.uservote_set.filter(locality__in=instance.locality.all()).distinct('user').count()


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    form = CategoryForm


@admin.register(UserVote)
class UserVoteAdmin(admin.ModelAdmin):
    list_display = [
        'esia_id',
        'vote',
        'locality',
    ]

    exclude = [
        'user',
    ]

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return False

    def esia_id(self, instance):
        return instance.user.esia_id


@admin.register(ImportXlsModel)
class ImportXlsAdmin(admin.ModelAdmin):
    pass


@admin.register(RejectReason)
class RejectReasonAdmin(admin.ModelAdmin):
    pass
