import io
import zipfile
from django.contrib import admin
from django.contrib.gis import forms
from django.http import HttpResponse
from openpyxl import Workbook

from modules.appeals.models import Appeal, DepartmentCategory


class AppealAdminForm(forms.ModelForm):
    gis_point = forms.PointField(
        widget=forms.OSMWidget(
            attrs={
                "default_lon": 92.86717,
                "default_lat": 56.01839,
                "default_zoom": 12,
            }
        )
    )


# TODO refactor and move this somewhere
def create_appeals_export_xlsx(appeals_queryset):
    wb = Workbook()

    ws = wb.active

    ws.column_dimensions["A"].width = 48

    rn = 1
    ws.cell(row=rn, column=1, value="Регистрационный номер")
    ws.cell(row=rn, column=2, value="Муниципальное образование")
    ws.cell(row=rn, column=3, value="Статус")
    ws.cell(row=rn, column=4, value="Дата поступления")
    ws.cell(row=rn, column=5, value="Дата передачи на рассмотрение")
    ws.cell(row=rn, column=6, value="Дата передачи в работу")
    ws.cell(row=rn, column=7, value="Дата предоставления ответа")
    ws.cell(row=rn, column=8, value="Категория")
    ws.cell(row=rn, column=9, value="Подкатегория")
    ws.cell(row=rn, column=10, value="Оператор")
    ws.cell(row=rn, column=11, value="Исполнитель")
    rn += 1

    for appeal in appeals_queryset:
        departments = DepartmentCategory.objects.filter(
            category=appeal.category
        ).distinct("department")
        department = None if len(departments) == 0 else departments[0].department
        ws.cell(row=rn, column=1, value=appeal.number)
        ws.cell(row=rn, column=2, value=appeal.locality.name)
        ws.cell(row=rn, column=3, value=appeal.state_name)
        ws.cell(row=rn, column=4, value=appeal.creation_date_time)
        ws.cell(row=rn, column=5, value=appeal.moderation_pass_date)
        ws.cell(row=rn, column=6, value=appeal.in_progress_begin_date)
        ws.cell(row=rn, column=7, value=appeal.responded_date)
        ws.cell(
            row=rn,
            column=8,
            value=appeal.category.parent.name if appeal.category.parent else "",
        )
        ws.cell(row=rn, column=9, value=appeal.category.name)
        ws.cell(row=rn, column=10, value=department.name if department else "")
        ws.cell(
            row=rn,
            column=11,
            value=";".join([e.name for e in appeal.contractors.all()]),
        )
        rn += 1

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    return xlsx_buffer


# TODO refactor and move this somewhere
def appeals_export(modeladmin, request, queryset):
    if queryset.count() == 0:
        return

    xlsx_buffer = create_appeals_export_xlsx(queryset.order_by("-creation_date_time"))
    filename = "appeals.xls"

    response = HttpResponse(content_type="application/excel")
    response["Content-Disposition"] = 'attachment; filename="{}"'.format(filename)
    response.write(xlsx_buffer.getvalue())

    return response


appeals_export.short_description = "Экспорт обращений"


@admin.register(Appeal)
class AppealAdmin(admin.ModelAdmin):
    # form = AppealAdminForm # todo
    list_display = [
        "number",
        "creation_date_time",
        "category",
        "locality",
        "is_public",
        "state",
    ]
    actions = [appeals_export]
