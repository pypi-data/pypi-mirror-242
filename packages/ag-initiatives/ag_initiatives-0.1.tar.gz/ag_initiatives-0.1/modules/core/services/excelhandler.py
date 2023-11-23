import os
import io
import sys
import random
import openpyxl as xls
from collections import OrderedDict
import openpyxl.styles as xls_styles
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.page import PageMargins

from django.utils import timezone
from rest_framework.exceptions import ValidationError

ALPHABET = "absdefghijklmnopqrstuvwxyz"


class Sheet:
    """
    Класс Sheet
    Используется для обработок параметров листов книги Excel
    Имеет встроенный счётчик и перечень экземпляров
    При создании получает поля:
        - объект листа книги Worksheet
        - заголовок листа
        - индекс листа
        - статус листа (активный/обычный)
    """

    count = 0
    instances = []

    def __init__(self):
        self.object = None
        self.title = None
        self.index = None
        self.status = None
        Sheet.count += 1
        Sheet.instances.append(self)

    def delete(self):
        if self in Sheet.instances:
            Sheet.count -= 1 if Sheet.count > 0 else 0
            Sheet.instances.pop(Sheet.instances.index(self))

    def __del__(self):
        if self in Sheet.instances:
            Sheet.count -= 1 if Sheet.count > 0 else 0
            Sheet.instances.pop(Sheet.instances.index(self))


class Cell:
    """
    Класс Cell
    Используется для обработок ячеек листа Excel
    Имеет встроенный счётчик и перечень экземпляров
    Имеет перечень букв латинского алфавита для именования колонок
    При создании получает поля:
        - объект ячейки
        - адрес ячейки с использованием латинского алфавита (может быть как адресом ячейки, так и диапазоном)
        - индекс строки
        - индекс колонки
        - значение ячейки
        - объект листа книги
        - формат ячейки
    """

    count = 0
    instances = []
    column_letters = "abcdefghijklmnopqrstuvwxyz"

    def __init__(self):
        self.object = None
        self.address = None
        self.row = None
        self.column = None
        self.value = None
        self.sheet = None
        self.number_format = None
        Cell.count += 1
        Cell.instances.append(self)

    def delete(self):
        if self in Cell.instances:
            Cell.count -= 1 if Cell.count > 0 else 0
            Cell.instances.pop(Cell.instances.index(self))

    def __del__(self):
        if self in Cell.instances:
            Cell.count -= 1 if Cell.count > 0 else 0
            Cell.instances.pop(Cell.instances.index(self))


class Column:
    """
    Класс Column
    Используется для обработок колонок листа
    Имеет встроенный счётчик и перечень экземпляров
    Имеет перечень букв латинского алфавита для именования колонок
    При создании получает поля:
        - объект колонки
        - латинское обозначение колонки
        - индекс колонки
        - ширина колонки
        - самая подходящая ширина
        - авто-размер
    """

    count = 0
    instances = []
    column_letters = "abcdefghijklmnopqrstuvwxyz"

    def __init__(self):
        self.object = None
        self.letter = None
        self.number = None
        self.width = None
        self.bestFit = None
        self.auto_size = None
        Column.count += 1
        Column.instances.append(self)

    def delete(self):
        if self in Column.instances:
            Column.count -= 1 if Column.count > 0 else 0
            Column.instances.pop(Column.instances.index(self))

    def __del__(self):
        if self in Column.instances:
            Column.count -= 1 if Column.count > 0 else 0
            Column.instances.pop(Column.instances.index(self))


class Row:
    """
    Класс Row
    Используется для обработок строк листа
    Имеет встроенный счётчик и перечень экземпляров
    При создании получает поля:
        - объект колонки
        - индекс строки
        - высота строки
    """

    count = 0
    instances = []

    def __init__(self):
        self.object = None
        self.number = None
        self.height = None
        Row.count += 1
        Row.instances.append(self)

    def delete(self):
        if self in Row.instances:
            Row.count -= 1 if Row.count > 0 else 0
            Row.instances.pop(Row.instances.index(self))

    def __del__(self):
        if self in Row.instances:
            Row.count -= 1 if Row.count > 0 else 0
            Row.instances.pop(Row.instances.index(self))


class Style:
    """
    Класс Style
    Используется для обработок стилей книги
    Имеет встроенный счётчик и перечень экземпляров
    При создании получает поля:
        - наименование стиля
        - шрифт:
            - название
            - размер
            - цвет
            - жирное начертание
            - курсивное начертание
        - заполненние ячейки:
            - цвет фона
            - цвет основной
            - тип заполнения (заливка)
        - граница:
            - стиль
            - цвет
            - расположение (слева, справа, сверху, снизу)
        - выравнивание:
            - горизонтальное
            - вертикальное
            - перенос текста
        - объект стиля
    Для работы со стилем имеются следующие методы:
        - delete - удалить объект стиля
        - get_style_by_name - получить объект стиля по названию (name)
        - set_alignment - установить выравнивание (horizontal, vertical, wrap_text):
            - horizontal:
                - general
                - left
                - center
                - right
                - fill
                - justify
                - centerContinuous
                - distributed
            - vertical:
                - top
                - center
                - bottom
                - justify
                - distributed
            - wrap_text:
                - True
                - False
        - set_border - установить границы
            - border_style:
                - dashDot
                - dashDotDot
                - dashed
                - dotted
                - double
                - hair
                - medium
                - mediumDashDot
                - mediumDashDotDot
                - mediumDashed
                - slantDashDot
                - thick
                - thin
            - border_color - 'rrggbb'
            - left - True/False
            - right - True/False
            - top - True/False
            - bottom - True/False
        - set_fill - установить заливку - bg_color='rrggbb'
        - set_font - установить шрифт:
            - name - наименование
            - size - размер
            - font_color - цвет 'rrggbb'
            - font_style - начертание (bold+italic)
        - set_name - установить название
    """

    count = 0
    instances = []

    def __init__(self):
        self.name = "Normal"
        self.font = xls_styles.Font(
            name="Liberation Serif",
            size=11.0,
            color="000000",
            bold=False,
            italic=False,
        )
        bgColor = xls_styles.Color("ffffff")
        fgColor = bgColor
        self.fill = xls_styles.PatternFill(
            bgColor=bgColor, fgColor=fgColor, fill_type="solid"
        )
        border_style = xls_styles.Side(style=None, color=None)
        self.border = xls_styles.Border(
            left=border_style, right=border_style, top=border_style, bottom=border_style
        )
        self.alignment = xls_styles.Alignment(
            horizontal="left", vertical="top", wrapText=False
        )
        self.object = xls_styles.NamedStyle(
            name=self.name,
            font=self.font,
            fill=self.fill,
            border=self.border,
            alignment=self.alignment,
        )
        Style.count += 1
        Style.instances.append(self)

    def delete(self):
        if self in Style.instances:
            Style.count -= 1 if Style.count > 0 else 0
            Style.instances.pop(Style.instances.index(self))

    @staticmethod
    def get_style_by_name(name):
        for style in Style.instances:
            if style.name == name:
                return style
        return None

    def set_alignment(self, **kwargs):
        horizontal = kwargs.get("horizontal", "left")
        vertical = kwargs.get("vertical", "top")
        wrap_text = kwargs.get("wrap_text", False)
        if horizontal not in [
            "general",
            "left",
            "center",
            "right",
            "fill",
            "justify",
            "centerContinuous",
            "distributed",
        ]:
            horizontal = "left"
        if vertical not in ["top", "center", "bottom", "justify", "distributed"]:
            vertical = "top"
        self.alignment = xls_styles.Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrapText=wrap_text,
            wrap_text=wrap_text,
        )
        self.object.alignment = self.alignment
        return self

    def set_border(self, **kwargs):
        style = kwargs.get("border_style", None)
        color = kwargs.get("border_color", None)
        _left_ = kwargs.get("left", True)
        _right_ = kwargs.get("right", True)
        _top_ = kwargs.get("top", True)
        _bottom_ = kwargs.get("bottom", True)
        if style not in [
            "dashDot",
            "dashDotDot",
            "dashed",
            "dotted",
            "double",
            "hair",
            "medium",
            "mediumDashDot",
            "mediumDashDotDot",
            "mediumDashed",
            "slantDashDot",
            "thick",
            "thin",
        ]:
            style = None
        default_border_style = xls_styles.Side(style=None, color=None)
        left = default_border_style
        right = default_border_style
        top = default_border_style
        bottom = default_border_style
        border_style = xls_styles.Side(style=style, color=color)
        if _left_:
            left = border_style
        if _right_:
            right = border_style
        if _top_:
            top = border_style
        if _bottom_:
            bottom = border_style
        self.border = xls_styles.Border(left=left, right=right, top=top, bottom=bottom)
        self.object.border = self.border
        return self

    def set_fill(self, **kwargs):
        bgColor = xls_styles.Color(kwargs.get("bg_color", "ffffff"))
        fgColor = bgColor
        self.fill = xls_styles.PatternFill(
            bgColor=bgColor, fgColor=fgColor, fill_type="solid"
        )
        self.object.fill = self.fill
        return self

    def set_font(self, **kwargs):
        name = kwargs.get("font_name", None)
        size = kwargs.get("font_size", None)
        color = kwargs.get("font_color", None)
        style = kwargs.get("font_style", None)
        italic = False
        bold = False
        if not str(size).isdigit():
            size = None
        else:
            size = float(size)
        if style and "italic" in str(style).lower():
            italic = True
        if style and "bold" in str(style).lower():
            bold = True

        self.font = xls_styles.Font(
            name=name, italic=italic, bold=bold, size=size, color=color
        )
        self.object.font = self.font
        return self

    def set_name(self, **kwargs):
        name = kwargs.get("name", "Normal")
        self.name = f"{name}"
        self.object.name = self.name
        return self

    def __del__(self):
        if self in Style.instances:
            Style.count -= 1 if Style.count > 0 else 0
            Style.instances.pop(Style.instances.index(self))


class Excel:
    """
    Внешняя адресация происходит по индексу колонок и строк:
        - колонки [0:1023]
        - строки [0:1048575]
    Внутренняя адресация происходит по номерам колонок и строк:
        - колонки [1:1024]
        - строки [1:1048576]
    Также внутренняя адресация работает через систему координат с
    использованием заглавных латинских букв в качестве имён столбцов

    worksheets = {
        <Worksheet NAME>: Sheet()
    }
    cells = {
        <Cell COORDINATE>: Cell()
    }
    rows = {
        <ROW>: Row()
    }
    columns = {
        <Column LETTER>: Column()
    }
    Создание объекта Excel:
        excel_obj = Excel()
        excel_obj.create() / excel_obj.open(filename)
    Работа с листами:
        excel_obj.create_sheet("Sheet2")
        excel_obj.checkout_sheet(title='Sheet2')
        excel_obj.pop_sheet(title='Sheet2')
    Работа с ячейками:
        excel_obj.edit(0, 0, <text>)
        excel_obj.set_height(0, 0, 50)
        excel_obj.edit(10, 5, 44484, 'dd-mm-yy')
        excel_obj.clear(0, 0)
        excel_obj.clear(2, 2, 4, 4)
        excel_obj.add_style(name='User3', border_style='thick', border_color='00FFFF', top=False, horizontal='center')
        excel_obj.add_style(name='Wrapped', wrap_text=True)
        excel_obj.set_style(21, 10, style='Wrapped')
        excel_obj.set_width(21, 6, 30)
    Сохранение объекта и чтение из памяти в ответ сервера
        excel_obj.save(in_memory=True)
        response.write(excel_obj.read_data_from_memory())
    """

    worksheets = OrderedDict()
    cells = OrderedDict()
    rows = OrderedDict()
    columns = OrderedDict()

    def __init__(self):
        self.today = timezone.now().astimezone()
        self.workbook = None
        self.file = io.BytesIO()
        self.style = None

    def _check_column(self, column):
        """Проверка вхождения колонки в границы листа"""
        if column and (column < 0 or column > 1023):
            column = 0
        return column

    def _check_row(self, row):
        """Проверка вхождения строки в границы листа"""
        if row and (row < 0 or row > 1048575):
            row = 0
        return row

    def _get_addresses(self, *args):
        """
        Получение адресов из аргументов:
            - вариант 1 - args = (start_row, start_column, end_row, end_column)
            - вариант 2 - args = (start_row, start_column)
        На основе входных данных формируются адреса:
            - минимальный - начальная позиция (минимальный столбец + минимальная строка)
            - максимальный - конечная позиция (максимальный столбец + максимальная строка) или None
            - общий - начальная позиция:конечная позиция или минимальная позиция
        """
        if len(args) == 4:
            range_boundaries = self._get_range_boundaries(*args)
            min_row, max_row, min_column, max_column = range_boundaries
        else:
            min_row, min_column = args
            max_row, max_column = None, None
        min_column = self._check_column(min_column)
        max_column = self._check_column(max_column)
        min_row = self._check_row(min_row)
        max_row = self._check_row(max_row)
        min_address = self._get_letter_cell_address(min_row, min_column)
        max_address = self._get_letter_cell_address(max_row, max_column)
        if max_address == min_address:
            max_address = None
        address = f"{min_address}:{max_address}" if max_address else f"{min_address}"
        return min_address, max_address, address

    def _get_letter_cell_address(self, row: int, column: int) -> str:
        """
        Формирование адреса ячейки из номера столбца [0: max_column-1] и номера строки [0: max_row-1]
        """
        address = None
        if not (row is None and column is None):
            if column < len(Cell.column_letters):
                letters = Cell.column_letters[column]
            elif (
                len(Cell.column_letters)
                <= column
                < len(Cell.column_letters) * (len(Cell.column_letters) + 1)
            ):
                k = len(Cell.column_letters)
                letters = (
                    f"{Cell.column_letters[(column - k) // k]}"
                    f"{Cell.column_letters[(column - k) % k]}"
                )
            elif (
                len(Cell.column_letters) * (len(Cell.column_letters) + 1)
                <= column
                <= (2**10) - 1
            ):
                k = len(Cell.column_letters) * (len(Cell.column_letters) + 1)
                letters = (
                    f"{Cell.column_letters[(column // k) - 1]}"
                    f"{Cell.column_letters[(column % k) // len(Cell.column_letters)]}"
                    f"{Cell.column_letters[(column % k) % len(Cell.column_letters)]}"
                )
            else:
                letters = "A"

            address = f"{letters}{row + 1}".upper()
        return address

    def _get_range_boundaries(self, *args):
        """Формирование начальных и конечных строк и столбцов диапазона"""
        start_row, start_column, end_row, end_column = args
        r1 = min(start_row, end_row)
        r2 = max(start_row, end_row)
        c1 = min(start_column, end_column)
        c2 = max(start_column, end_column)
        return r1, r2, c1, c2

    def _set_sheet(
        self,
        sheet: Sheet,
        object=None,
        title: str = None,
        index: int = None,
        sheet_status: str = None,
    ):
        """Установка параметров листа"""
        sheet.object = object
        sheet.title = title
        sheet.index = index
        sheet.status = sheet_status
        self.worksheets[sheet.object] = sheet

    def add_style(self, **kwargs):
        """Добавление стиля в книгу"""
        style = Style()
        style.set_name(**kwargs)
        style.set_font(**kwargs)
        style.set_fill(**kwargs)
        style.set_border(**kwargs)
        style.set_alignment(**kwargs)
        if style.name not in self.workbook.named_styles:
            self.workbook.add_named_style(style.object)
        else:
            raise Exception(f"Style {style.name} already exists")

        return self

    def checkout_sheet(self, **kwargs):
        """Переключение на лист по индексу или наименованию"""
        sheet_identifier = kwargs.get("index", kwargs.get("title", None))
        new_active_sheet = None
        if sheet_identifier and str(sheet_identifier).isdigit():
            try:
                sheet_name = self.get_sheet_list()[int(sheet_identifier)]
                new_active_sheet = self.workbook[sheet_name]
            except Exception as e:
                pass
            else:
                self.active_sheet = new_active_sheet
        elif sheet_identifier:
            try:
                new_active_sheet = self.workbook[sheet_identifier]
            except Exception as e:
                pass
            else:
                self.active_sheet = new_active_sheet

        if sheet_identifier and new_active_sheet:
            for sheet_data in self.worksheets.values():
                sheet_data.status = "usual"
            sheet = self.worksheets.get(self.active_sheet, None)
            if not sheet:
                sheet = Sheet()
            self._set_sheet(
                sheet,
                self.active_sheet,
                self.active_sheet.title,
                self.get_sheet_list().index(self.active_sheet.title),
                "active",
            )
            self.workbook.active = self.active_sheet
        return self

    def clear(self, *args):
        """Очистка ячейки/диапазона от данных"""
        min_address, max_address, address = self._get_addresses(*args)

        cells_for_clear = []
        for cell in self.cells.values():
            if cell.address == address and cell.sheet == self.active_sheet:
                if ":" in cell.address:
                    self.active_sheet.unmerge_cells(cell.address)
                cell.object.value = ""
                cells_for_clear.append(cell)
        for cell in cells_for_clear:
            if cell.object.coordinate in self.cells:
                self.cells.pop(cell.object.coordinate)
                cell.delete()
        cells_for_clear = None
        return self

    def clear_style(self, *args):
        """Сброс стиля ячейки на стиль По умолчанию"""
        min_address, max_address, address = self._get_addresses(*args)
        self.active_sheet[min_address].style = self.default_style.name
        if max_address:
            self.active_sheet.merge_cells(address)

    def create(self, title: str = None):
        """
        Создание объекта Excel
        Задания заголовка листа через title, иначе имя по умолчанию
        """
        self.workbook = xls.Workbook()
        self.active_sheet = self.workbook.active
        self.set_default_style()
        if title:
            self.active_sheet.title = f"{title}"
        sheet = Sheet()
        self._set_sheet(
            sheet,
            self.active_sheet,
            self.active_sheet.title,
            self.get_sheet_list().index(self.active_sheet.title),
            "active",
        )

    def create_sheet(self, title: str = None, index: int = None):
        """
        Создание листов книги Excel
        На вход:
            - title - заголовок листа
            - index - индекс листа
        """
        if self.workbook:
            if title and index:
                ws = self.workbook.create_sheet(f"{title}", index)
            elif title:
                ws = self.workbook.create_sheet(f"{title}")
            else:
                ws = self.workbook.create_sheet()
                title = ws.title
            sheet = Sheet()
            self._set_sheet(
                sheet, ws, title, self.get_sheet_list().index(ws.title), "usual"
            )
        else:
            raise Exception("Workbook does not exist")
        return self

    def edit(self, *args, **kwargs):
        """
        Редактирование ячейки/диапазона
        На вход:
            - вариант 1 - args = (start_row, start_column, end_row, end_column, value)
            - вариант 2 - args = (start_row, start_column, end_row, end_column, value, value_type)
            - вариант 3 - args = (start_row, start_column, value)
            - вариант 4 - args = (start_row, start_column, value, value_type)
        Список доступных типов можно получить через Excel.get_number_formats()

        """
        if len(args) == 3 or len(args) == 5:
            value = args[-1]
            value_type = self.get_number_formats(0)
            args = args[:-1]
        else:
            value = args[-2]
            if str(args[-1]).isdigit() and int(args[-1]) in self.get_number_formats():
                value_type = self.get_number_formats(int(args[-1]))
            elif isinstance(args[-1], str):
                value_type = f"{args[-1]}"
            else:
                value_type = self.get_number_formats(0)
            args = args[:-2]
        min_address, max_address, address = self._get_addresses(*args)
        self.get_number_formats()
        if max_address:
            self.active_sheet.merge_cells(address)
        cell = Cell()
        cell.object = self.active_sheet[min_address]
        if "hyper" in kwargs or "hyperlink" in kwargs or "link" in kwargs:
            hyperlink = kwargs.get(
                "hyper", kwargs.get("hyperlink", kwargs.get("link", ""))
            )
            cell.object.hyperlink = f"{hyperlink}"
            cell.object.style = "Hyperlink"
        cell.object.value = value

        cell.object.number_format = value_type
        cell.address = address
        cell.row = self.active_sheet[min_address].row
        cell.column = self.active_sheet[min_address].column
        cell.value = value
        cell.number_format = value_type
        cell.sheet = self.active_sheet
        self.cells[cell.object.coordinate] = cell
        return self

    def get_sheet_list(self):
        """Возвращает список листов книги"""
        return self.workbook.sheetnames

    def open(self, file, read_only: bool = True):
        """Открытие файла Excel"""
        if os.path.exists(file) and os.path.isfile(file):
            self.path = os.path.split(file)[0]
            self.filename = os.path.splitext(os.path.split(file)[-1])[0]
            self.filetype = os.path.splitext(os.path.split(file)[-1])[-1]
            if ".xls" in self.filetype:
                try:
                    self.workbook = xls.open(file, read_only=read_only)
                except Exception as e:
                    raise ValidationError(e.args, type(e))
        return self

    def pop_sheet(self, **kwargs):
        """Удаление листа из книги"""
        sheet_identifier = kwargs.get("index", kwargs.get("title", None))
        pop_sheet = None
        if sheet_identifier and str(sheet_identifier).isdigit():
            try:
                sheet_name = self.get_sheet_list()[int(sheet_identifier)]
                pop_sheet = self.workbook[sheet_name]
            except Exception as e:
                pass
        elif sheet_identifier:
            try:
                pop_sheet = self.workbook[sheet_identifier]
            except Exception as e:
                pass

        if sheet_identifier and pop_sheet:
            active_conditions = [
                pop_sheet in self.worksheets,
                self.worksheets[pop_sheet].status is "active",
                len(self.worksheets) >= 2,
            ]
            if all(active_conditions):
                try:
                    new_active_sheet = self.workbook[
                        self.get_sheet_list()[self.worksheets[pop_sheet].index - 1]
                    ]
                except Exception as e:
                    pass
                else:
                    self.active_sheet = new_active_sheet
                    self.workbook.active = self.active_sheet
                    self.worksheets[self.active_sheet].status = "active"

            pop_sheet = self.worksheets.get(pop_sheet)
            self.worksheets.pop(pop_sheet.object)
            self.workbook.remove_sheet(pop_sheet.object)
            pop_sheet.delete()
            for index, sheet_title in enumerate(self.get_sheet_list()):
                for sheet, sheet_data in self.worksheets.items():
                    if sheet_title == sheet_data.title:
                        self.worksheets[sheet].index = index

        return self

    def read_data_from_memory(self):
        """Чтение книги из памяти"""
        if self.file:
            return self.file.read()

    def rename_sheet(self, title: str):
        """Переименование активного листа"""
        if self.workbook and self.active_sheet:
            if self.active_sheet in self.worksheets:
                self.worksheets[self.active_sheet].title = f"{title}"
            else:
                sheet = Sheet()
                self._set_sheet(
                    sheet,
                    self.active_sheet,
                    self.active_sheet.title,
                    self.get_sheet_list().index(self.active_sheet.title),
                    "active",
                )

            self.active_sheet.title = f"{title}"
        return self

    def save(self, file: str = None, in_memory=True):
        """Сохранение книги в файл или в память"""
        base_file = "".join(
            [
                [random.choice(ALPHABET), random.choice(ALPHABET).upper()][
                    random.randint(0, 1)
                ]
                for _ in range(10)
            ]
        )
        self.file = f'{base_file}_{self.today.strftime("%Y-%m-%d %H:%M:%S")}'
        if file:
            self.file = file
        if ".xlsx" not in self.file:
            self.file += ".xlsx"
        if not os.path.exists(self.file):
            self.file = io.BytesIO()

        self.workbook.save(self.file)
        if in_memory:
            self.file.seek(0)
        return self

    def set_default_style(self):
        """Формирование стиля По умолчанию"""
        self.default_style = Style()
        self.default_style.name = self.workbook.style_names[-1]
        self.default_style.fill = self.active_sheet["A1"].fill
        self.default_style.font = self.active_sheet["A1"].font
        self.default_style.border = self.active_sheet["A1"].border
        self.default_style.alignment = self.active_sheet["A1"].alignment
        self.default_style.object = self.active_sheet["A1"].style
        return self

    def set_height(self, *args, **kwargs):
        """
        Установка высоты строки
        На вход:
            - вариант 1 - args = (start_row, start_column, end_row, end_column)
            - вариант 2 - args = (start_row, start_column)
            - вариант 3 - args = (start_row, start_column, end_row, end_column, height)
            - вариант 4 - args = (start_row, start_column, height)
        Если значение height отсутствует, происходит попытка установить высоту автоматически,
        в зависимости от содержимого ячеек строки
        """
        if len(args) == 3 or len(args) == 5:
            height = args[-1]
            args = args[:-1]
        else:
            height = None

        if kwargs.get("height", None):
            height = kwargs.get("height")
        if height is None and kwargs.get("height", None) is None:
            height = None

        if height and height > 409:
            height = 409
        if height and height < 0:
            height = 0
        min_address, max_address, address = self._get_addresses(*args)

        row = Row()
        row.number = int(self.get_cell_row(min_address))
        row.object = self.active_sheet.row_dimensions[row.number]
        row.height = row.object.height
        self.columns[row.number] = row
        row_cells = []
        cell_columns = []
        if not height:
            for cell in self.cells.values():
                if cell.row == row.number and cell not in row_cells:
                    row_cells.append(cell)
            for cell in row_cells:
                column_letter = self.get_cell_column(cell.address)
                if column_letter in self.columns:
                    column = self.columns[self.get_cell_column(cell.address)]
                else:
                    column = Column()
                    column.letter = self.get_cell_column(min_address)
                    column.number = int(min_address.lstrip(column.letter)) - 1
                    self.columns[column.letter] = column
                cell_columns.append(column)
                if not column.width:
                    self.set_width(*args)
                if cell.value and column.width:
                    new_height = (
                        (len(cell.value) * 1.23) / column.width
                    ) * 15.0568181818182
                    if not height or new_height > height:
                        self.set_width(*args)
                        height = new_height

        if height:
            row.object.height = height
            row.height = height

        return self

    def set_page_orientation(self, orientation: str):
        if str(orientation).lower() == "portrait":
            self.active_sheet.page_setup.orientation = (
                self.active_sheet.ORIENTATION_PORTRAIT
            )
        elif str(orientation).lower() == "landscape":
            self.active_sheet.page_setup.orientation = (
                self.active_sheet.ORIENTATION_LANDSCAPE
            )
        return self

    def set_style(self, *args, style: str = None):
        """Установка существующего стиля для ячейки/диапазона"""
        min_address, max_address, address = self._get_addresses(*args)
        self.active_sheet[min_address].style = style

        # self.active_sheet[min_address].alignment.wrap_text = st.alignment.wrap_text
        if max_address:
            self.active_sheet.merge_cells(address)

        return self

    def set_width(self, *args, **kwargs):
        """
        Установка ширины столбца
        На вход:
            - вариант 1 - args = (start_row, start_column, end_row, end_column)
            - вариант 2 - args = (start_row, start_column)
            - вариант 3 - args = (start_row, start_column, end_row, end_column, width)
            - вариант 4 - args = (start_row, start_column, width)
        Если значение width отсутствует, происходит попытка установить ширину автоматически
        """
        if len(args) == 3 or len(args) == 5:
            width = args[-1]
            args = args[:-1]
        else:
            width = None

        if kwargs.get("width", None):
            width = kwargs.get("width")
        if width is None and kwargs.get("width", None) is None:
            width = None

        if width and width > 255:
            width = 255
        if width and width < 0:
            width = 0
        min_address, max_address, address = self._get_addresses(*args)
        column = Column()
        column.letter = self.get_cell_column(min_address)
        column.number = int(min_address.lstrip(column.letter)) - 1
        column.object = self.active_sheet.column_dimensions[column.letter]
        column.width = column.object.width
        column.object.bestFit = True
        column.object.auto_size = True
        column.bestFit = column.object.bestFit
        column.auto_size = column.object.auto_size
        self.columns[column.letter] = column
        if width:
            column.object.width = width
            column.width = width

        return self

    def update_style(self, name: str, **kwargs):
        style = None
        for item in Style.instances:
            if item.name == name:
                style = item
                break
        if style:
            style.set_font(**kwargs)
            style.set_fill(**kwargs)
            style.set_border(**kwargs)
            style.set_alignment(**kwargs)

        return self

    @staticmethod
    def get_cell_column(address: str) -> str:
        """Получение имени столбца из адреса"""
        return "".join(
            [item for item in address.split(":")[0] if not str(item).isdigit()]
        )

    @staticmethod
    def get_cell_row(address: str) -> int:
        """Получение номера строки из адреса"""
        return int(
            "".join([item for item in address.split(":")[0] if str(item).isdigit()])
        )

    @staticmethod
    def get_number_formats(key: int = None):
        """
        Получение списка доступных форматов ячейки
        Получение формата ячейки по индексу [0:49]
        """
        number_formats = xls_styles.numbers.BUILTIN_FORMATS.copy()
        if key is None:
            for key in sorted(number_formats.keys()):
                pass
            return number_formats
        else:
            return number_formats.get(key, number_formats[0])
