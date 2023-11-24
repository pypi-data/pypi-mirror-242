# -*- encoding: utf-8 -*-
"""
@File    :   read.py
@Time    :   2023-02-17 23:02
@Author  :   坐公交也用券
@Version :   1.0
@Contact :   faith01238@hotmail.com
@Homepage : https://liumou.site
@Desc    :   使用Openpyxl模块实现的读取功能
"""
from sys import exit

from loguru import logger
from openpyxl import load_workbook

from .dic import GetDic
from .exists import _exists


class Read:
	def __init__(self, filename):
		"""
		读取表格数据
		:param filename: 需要读取的表格文件
		"""
		self.logger = logger
		self.Err = None  # 设置错误信息

		self.debug = False
		self.filename = filename
		if not _exists(filename):
			self.logger.error(f"文件异常: {self.filename}")
			exit(1)
		# 实例
		self.wb = load_workbook(filename=self.filename)

		# 获取信息
		self.InfoSheetList = self.wb.sheetnames  # 获取所有sheet名称
		self.InfoSheetName = None  # 存储工作簿名称
		self.InfoSheet = None

		self.ws = self.wb.active
		self.InfoRows = self.ws.max_row  # 获取行数
		self.InfoCols = self.ws.max_column  # 获取列数

		# 设置数据变量
		self.DataR = []  # 首次读取的数据
		self.Data = []  # 最终处理的数据
		# 拆分数据
		self.DataSplit = []
		# 设置数据常量
		self._Dic = GetDic()
		self.set()  # 读取默认配置

	def set(self, debug=False, sheet_name=None, sheet_index=None):
		"""
		自定义设置
		:param debug: 是否开启Debug
		:param sheet_name: 通过工作簿名称选择,当设置了此项切勿再设置索引(默认通过索引设置/当同时设置则优先使用索引值)
		:param sheet_index: 通过工作簿索引值选择(默认:0)
		:return:
		"""
		self.debug = debug
		# 判断读取方式
		if sheet_name is None and sheet_index is None:
			self.InfoSheet = 0
		else:
			if sheet_index and sheet_name:
				self.InfoSheet = sheet_index
			else:
				if sheet_index is None:
					if sheet_name in self.InfoSheetList:
						self.InfoSheet = self.InfoSheetList.index(sheet_name)
					else:
						logger.error(f"找不到sheet名称: {sheet_name}")
						exit(2)
				else:
					self.InfoSheet = sheet_index
		self.InfoSheetName = self.InfoSheetList[self.InfoSheet]
		self.wb.get_sheet_by_name(self.InfoSheetName)  # 指定工作簿

	def base(self):
		"""
		打印表格文件基础信息
		:return:
		"""
		self.logger.debug(f"当前文件: {self.filename}")
		self.logger.debug(f"当前sheet列表: {self.InfoSheetList}")
		self.logger.debug(f"当前选择sheet索引: {self.InfoSheet}")
		self.logger.debug(f"当前sheet总行数: {self.InfoRows}")
		self.logger.debug(f"当前sheet总列数: {self.InfoCols}")

	def read_line(self, row=1):
		"""
		读取某行数据
		:param row: 需要读取的行
		:return:
		"""
		data_ = []
		for col in range(self.InfoCols):
			col += 1
			if self.debug:
				self.logger.debug(f"正在读取行: {row}")
				self.logger.debug(f"正在读取列: {col}")
			value = self.ws.cell(row=row, column=col).value
			data_.append(value)
		return data_

	def read_col(self, col=1):
		"""
		读取某列所有数据
		:param col: 需要读取的列
		:return:
		"""
		data_ = []
		for row in range(self.InfoRows):
			row += 1
			if self.debug:
				self.logger.debug(f"正在读取行: {row}")
				self.logger.debug(f"正在读取列: {col}")
			value = self.ws.cell(row=row, column=col).value
			data_.append(value)
		return data_

	def read_all(self, index=True):
		"""
		获取指定Sheet所有数据,可通过Set函数设置指定Sheet
		:param index: 是否读取首行,默认读取
		:return:
		"""
		data_ = []
		r = 1
		if index:
			r = 0
		for row in range(self.InfoRows):
			r += 1
			row_list = []
			for clo in range(self.InfoCols):
				clo += 1
				value = self.ws.cell(row=r, column=clo).value
				row_list.append(value)
			data_.append(row_list)
		self.Data = data_
		self.DataR = data_
		return self

	def _check(self, start, end, data):
		"""
		检查传入的数值是否符合实际要求
		:param start: 开始数
		:param end: 截止数
		:param data: 总数
		:return: bool
		"""
		if int(start) >= int(end):
			self.Err = "开始数大于截止数"
			return False
		if int(end) > int(data):
			self.Err = "截止数大于实际数"
			return False
		if int(data) <= int(start):
			self.Err = "实际数小于开始数"
			return False
		return True

	def split_data(self, copies=1):
		"""
		对已处理的数据进行拆分
		:param copies: 需要拆分多少份,如果无法整除,那么最终份数会比设置的份数多1份
		:return: self.DataSplit
		"""
		# 获取总行数
		row = len(self.Data)
		if copies == 1:
			logger.warning(f"未进行数据拆分...")
			self.DataSplit = self.Data
			return
		# 获取每一份的数量
		number = int(row) // int(copies)
		# 获取最后一份的数量
		number_end = int(row) - int(number) * int(copies)
		logger.info(f"根据参数对当前数据拆分成: {copies} 份, 每份数量: {number} ,其中最后一份数量: {number_end}")
		data = []
		all_ = 1
		sum_ = 1
		for i in self.Data:
			data.append(i)
			if int(sum_) == number:
				self.DataSplit.append(data)
				data = []
				sum_ = 0
			else:
				if all_ == int(row):
					self.DataSplit.append(data)
					data = []
					sum_ = 0
			sum_ += 1
			all_ += 1
		return self.DataSplit

	def read_line_range(self, start=0, end=1):
		"""
		读取行范围(1是第一行)
		:param start: 起始行
		:param end: 结束行
		:return:获取结果(bool),获取数据请通过实例变量(DataR)获取
		"""
		self.Err = None
		data_ = []
		end += 1
		if self._check(start=start, end=end, data=self.InfoRows):
			for row in range(start, end):
				row_list = []
				for clo in range(self.InfoCols):
					clo += 1
					value = self.ws.cell(row=row, column=clo).value
					row_list.append(value)
				data_.append(row_list)
			self.Data = data_
			self.DataR = data_

		return self

	def read_column_range(self, start=0, end=1):
		"""
		读取列范围
		:param start: 起始列
		:param end: 结束列
		:return: 获取结果(bool),获取数据请通过实例变量(DataR)获取
		"""
		self.Err = None
		data_ = []
		if self._check(start=start, end=end, data=self.InfoRows):
			for row in range(self.InfoRows):
				row += 1
				row_list = []
				for clo in range(start, end):
					value = self.ws.cell(row=row, column=clo).value
					row_list.append(value)
					clo += 1
				data_.append(row_list)
			self.Data = data_
			self.DataR = data_

		return self

	def cut_line(self, n=0):
		"""
		对已读取的数据进行截取指定行
		:param n: 需要读取的行
		:return: 取结果(bool),获取数据请通过实例变量(DataR)获取
		"""
		if self.Err is None:
			if len(self.DataR) >= n:
				self.DataR = self.DataR[n]
			else:
				self.Err = "行数据不足,无法截取"
				self.logger.error(self.Err)
		else:
			self.logger.error(f"已存在错误,请先处理: {self.Err}")
		return self

	def cut_column(self, col=0):
		"""
		对已读取的数据进行列截取
		:param col: 需要截取的列
		:return:
		"""
		data_list = []  # 创建一个临时变量存储处理数据
		line = 0
		for i in self.DataR:
			if len(i) >= col:
				data_list.append(i[col])
			else:
				self.logger.warning(f"第 {str(line)} 行列数量不足,无法切割")
				self.logger.debug(i)
			line += 1
		self.DataR = data_list  # 将临时变量赋予实例变量
		return self

	def cut_line_range(self, start=0, end=1):
		"""
		对已读取的数据进行 行范围 截取
		:param start: 起始行
		:param end: 结束行
		:return: 取结果(bool),获取数据请通过实例变量(DataR)获取
		"""
		if start < end:
			if self.Err is None:
				if len(self.DataR) >= end:
					self.DataR = self.DataR[start:end]
				else:
					self.Err = "行数据不足,无法截取"
					self.logger.error(self.Err)
			else:
				self.logger.error(f"已存在错误,请先处理: {self.Err}")
		else:
			self.Err = "开始行数大于或大于结束行数"
			self.logger.error(self.Err)
		return self

	def cut_column_range(self, start=0, end=1):
		"""
		对已读取的数据进行 列范围 截取
		:param start: 起始列
		:param end: 结束列
		:return:
		"""
		dataList = []  # 创建一个临时变量存储处理数据
		line = 0
		if start < end:
			for i in self.DataR:
				if len(i) >= end:
					dataList.append(i[start:end])
				else:
					self.logger.warning(f"第 {line} 行列数量不足,无法切割")
					self.logger.debug(i)
				line += 1
			self.DataR = dataList  # 将临时变量赋予实例变量
		else:
			self.Err = "设置开始列数大于或等于结束列数"
		return self
