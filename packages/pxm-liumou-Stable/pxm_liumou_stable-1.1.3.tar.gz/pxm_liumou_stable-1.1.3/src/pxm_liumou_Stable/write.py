# -*- encoding: utf-8 -*-
"""
@File    :   write.py
@Time    :   2023-02-17 23:02
@Author  :   坐公交也用券
@Version :   1.0
@Contact :   faith01238@hotmail.com
@Homepage : https://liumou.site
@Desc    :   使用Openpyxl模块实现的写入功能
"""
from sys import exit

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from .exists import _exists


class Write:
	def __init__(self, filename, create=True):
		"""
		写入表格数据
		:param filename: 需要写入的表格文件
		:param create: 是否自动创建文件
		"""
		# 文件不存在是否自动创建
		self.auto_create_xlsx = create
		self.logger = logger
		self.filename = filename
		if _exists(filename):
			logger.debug(f"使用加载文件的方式写入内容: {self.filename}")
			self.wb = load_workbook(filename=self.filename)
		else:
			if self.auto_create_xlsx:
				logger.debug(f"使用创建文件的方式写入内容: {self.filename}")
				self.wb = Workbook()
			else:
				logger.error(f"表格文件不存在且已关闭自动创建文件选项: {self.filename}")
				exit(2)
		self.ws = self.wb.active
		self.SetSerialNumber = False  # 是否自动添加序号(默认False)
		self.SetAdd = True
		self.SetHead = True
		self.SetLine = None
		self.SetSpeedSum = 100
		self._WrData = []  # 设置需要写入的数据列表
		self.Err = None

	def create_sheet(self, name, index=None):
		"""
		创建sheet
		:param name: 新sheet的名称
		:param index: 设置sheet排序位置(0是排第一)
		:return: 创建结果
		"""
		if index is None:
			self.wb.create_sheet(title=name)
		else:
			self.wb.create_sheet(title=name, index=index)
		self.wb.save(self.filename)

	def update_line(self, row, data_list):
		"""
		更新某行数据
		:param row: 需要更新的行
		:param data_list: 行数据列表
		:return:
		"""
		try:
			col = 0
			for i in data_list:
				self.ws.cell(row=row, column=col, value=i)
				col += 1
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def write_cover_lists(self, lists, index=False):
		"""
		通过列表方式覆盖写入数据,一次性最多写入104万行数据
		:param index: 是否保留首行
		:param lists: 写入数据列表,例如: [["张三", "男", "33"], ["李四", "男", "32"]]
		:return:
		"""
		s = 1
		if index:
			s = 2
		self.ws.delete_rows(s, self.ws.max_row)
		self.wb.save(self.filename)
		try:
			for i in lists:
				self.ws.append(i)
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def write_lists(self, lists):
		"""
		通过列表方式追加写入数据,一次性最多写入104万行数据
		:param lists: 写入数据列表,例如: [["张三", "男", "33"], ["李四", "男", "32"]]
		:return:
		"""
		try:
			for i in lists:
				self.ws.append(i)
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def write_add_line(self, data):
		"""
		追加写入一行数据
		:param data: 数据,以列表形式 ["张三", "男", "33"]
		或者字典模式1: {"A": "刘某", "B": "男", "C": "22"}
		字典模式2: {1: 1, 2: 2, 3: 3}
		:return:
		"""
		try:
			self.ws.append(data)
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def write_add_col(self, col, data):
		"""
		写入一列数据
		:param col: 列数
		:param data: 数据
		:return:
		"""
		column = self.ws[col]
		try:
			for i in range(len(data)):
				column[i + 1].value = data[i]
			self.ws.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def delete_line(self, index, row=1):
		"""
		删除行数据
		:param index: 需要删除的起始行
		:param row: 需要删除的行数总数,默认删除1行(也就是起始行)
		:return: 删除结果
		"""
		self.Err = None
		try:
			self.ws.delete_rows(idx=index, amount=row)
			self.wb.save(self.filename)
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
		return self

	def set(self, add=True, head=True, line=None, speed_sum=100, serial_number=False):
		"""
		设置写入参数
		:param add: 是否使用追加模式(默认: True)
		:param head: 是否保留表头标题(默认: True)
		:param line: 是否自定义写入的行，如果需要自定义，请传所在行的整数
		:param speed_sum: 写入多少行数据进行一次进度显示
		:param serial_number: 是否自动添加序号(默认False)
		:return:
		"""
		self.SetSerialNumber = serial_number
		self.SetAdd = add
		self.SetHead = head
		self.SetLine = line
		if not self.SetAdd:
			# 如果使用覆盖模式,则设置初始行: 0
			self.SetLine = 0
		self.SetSpeedSum = speed_sum

	def _center_whole(self):
		"""
		设置全局居中
		:return:
		"""
		logger.info("正在设置全局居中")
		align = Alignment(horizontal='center', vertical='center')
		for row in self.ws.iter_rows():
			for cell in row:
				cell.alignment = align

	def _center_line(self, line: int):
		"""
		设置指定行居中
		:param line: 需要居中的行
		:return:
		"""
		logger.info(f"正在设置[ {line} ]行居中")
		align = Alignment(horizontal='center')
		first_row = self.ws[line]
		for cell in first_row:
			cell.alignment = align

	def _center_col(self, col: int):
		"""
		设置指定列居中
		:param col: 需要居中的列
		:return:
		"""
		align = Alignment(horizontal='center')
		first_column = [cell for row in self.ws for cell in row if cell.column == col]
		for cell in first_column:
			cell.alignment = align

	def set_center(self, whole=True, line=None, col=None, lines=None, cols=None):
		"""
		设置居中,此功能会遍历所有参数进行居中设置(如果设置了全局则忽略后面的参数)
		:param whole: 全局居中(bool)
		:param line: 指定行居中(int)
		:param col: 指定列居中(int)
		:param lines: 指定行范围居中,通过列表传入需要居中的行([int, int])
		:param cols: 指定列范围居中,通过列表传入需要居中的列([int, int])
		:return:
		"""
		if whole:
			self._center_whole()
			return
		if line is not None:
			self._center_line(line=line)
		if col is not None:
			self._center_col(col)
		if cols is not None:
			for c in cols:
				self._center_col(col=c)
		if lines is not None:
			for row in lines:
				self._center_line(line=row)
