#! /usr/bin/env python
# -*- coding:Utf-8 -*-
# pylint: disable=fixme,invalid-name,line-too-long

"""Excel system management."""
import math

# ============================================================
#    Linux python path and Library import
# ============================================================
import os.path

try:
    import win32com
    import win32com.client
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Border, Side
    from openpyxl.utils import cols_from_range, rows_from_range
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter

    # from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.cell_range import CellRange

    if win32com.client.gencache.is_readonly is True:
        # delete the directory Python\...\Lib\site-packages\win32com\gen_py
        # allow gencache to create the cached wrapper objects
        win32com.client.gencache.is_readonly = False
        # under p2exe the call in gencache to __init__() does not happen
        # so we use Rebuild() to force the creation of the gen_py folder
        win32com.client.gencache.Rebuild()
    import datetime

    from win32com import client as com
    from win32com.client import constants as com_const

    from . import libLog, libTool
except ImportError:
    # pylint: disable-next=raise-missing-from,broad-exception-raised
    raise Exception(
        """Could not load libExcel due to import Error
        Are you sure to installed Extras : pip install mecapacktools[Excel] ?"""
    )

# ============================================================
#    Class
# ============================================================


# ||||||||||||||||||||||||||||||||||||||||||||||||||
#    Excel
# ||||||||||||||||||||||||||||||||||||||||||||||||||
class Excel:
    """Class to manage the excel system."""

    # //////////////////////////////////////////////////
    #    Variables and Constants
    # //////////////////////////////////////////////////

    #: Variables
    hshParam = {}
    hshParam["visible"] = False

    # //////////////////////////////////////////////////
    #     INITIALIZATION
    # //////////////////////////////////////////////////
    def __init__(self, phshParam={}):  # pylint: disable=dangerous-default-value
        # Work variables
        try:
            self.__ApplicationXL = com.DispatchEx("Excel.Application")
            com.gencache.EnsureDispatch("Excel.Application")
        except Exception as e:
            # pylint: disable-next=raise-missing-from,broad-exception-raised
            raise Exception(
                f"error : {e} \n\n please check folder {win32com.__gen_path__}\n ther might be a problem with this"
            )
        self.constants = com_const
        self.__TemplateWB = object
        self.CurrentRange_Parts = {"ALL": None, "HEADER": None, "DATA": None}

        # Start log manager
        self.log = libLog.Log()

        # Update of parameters
        self.hshParam.update(phshParam)

    # //////////////////////////////////////////////////
    #     TemplateWB_Load
    # //////////////////////////////////////////////////
    def TemplateWB_Load(self, ptemplateWB_path):
        """
        Loading template workbook

        Args:
            ptemplateWB_path (_type_): the excel template path

        Raises:
            self.log.CustomException: Template file does not exist
        """

        # Visible
        self.__ApplicationXL.Visible = self.hshParam["visible"]

        # Optimization of load
        self.__ApplicationXL.ScreenUpdating = self.hshParam["visible"]
        self.__ApplicationXL.DisplayStatusBar = self.hshParam["visible"]
        self.__ApplicationXL.EnableEvents = self.hshParam["visible"]

        # open workbook template
        if os.path.isfile(ptemplateWB_path):
            self.__TemplateWB = self.__ApplicationXL.Workbooks.Open(ptemplateWB_path, ReadOnly=True)
        else:
            raise self.log.CustomException(f"Template file does not exist :{ptemplateWB_path}")

    # //////////////////////////////////////////////////
    #     TemplateWB_LoadData
    # //////////////////////////////////////////////////
    def TemplateWB_LoadData(self, psheet_name, pranges_name, pcursrows, pcurscolumns=(), **kw):
        """
        Loading data in excel template

        Args:
            psheet_name (str):  the name of target sheet
            pranges_name (str):  the name of target ranges
            pcursrows (rows): Rows data of cursor
            pcurscolumns (tuple, optional): Columns of cursor. Defaults to ().
            **prowindex -- Start index of row (default=0)
            **pcolindex -- Start index of column (default=0)
            **pdataclean -- Cleaning of old data (default=True)
            **pcursorloadmaxrows -- Cursor to loading : Max rows (default=100000)
            **preferstodata -- New affectation of named range with data size (default=False)
        """
        hshOption = {
            "prowindex": 0,
            "pcolindex": 0,
            "pdataclean": True,
            "pcursorloadmaxrows": 100000,
            "preferstodata": False,
        }
        self.CurrentRange_Parts["HEADER"] = None
        self.CurrentRange_Parts["DATA"] = None
        self.CurrentRange_Parts["ALL"] = None

        # Setting dictionary option
        if isinstance(kw, dict):
            hshOption.update(kw)

        # Select sheet
        sheetTemp = self.__TemplateWB.Worksheets(psheet_name)
        # Disable automatic calculation : optimization of insertion data
        sheetTemp.EnableCalculation = False
        # Define Range
        r = sheetTemp.Range(pranges_name)
        rangeTemp = sheetTemp.Range(
            r.Cells(1 + hshOption["prowindex"], 1 + hshOption["pcolindex"]),
            r.Cells(r.Rows.Count, r.Columns.Count),
        )
        # Delete old data
        if hshOption["pdataclean"]:
            rangeTemp.Value = None
        # Define Start index
        iColumn_Start = rangeTemp.Column
        iRow_Start = rangeTemp.Row
        iRow_Start2 = rangeTemp.Row
        # Define range counter
        iRange_ColumnCount = rangeTemp.Columns.Count
        iRange_RowCount = rangeTemp.Rows.Count
        # Define data counter
        iData_ColumnCount = len(pcurscolumns)
        if iData_ColumnCount == 0:
            iData_Column = False
        else:
            iData_Column = True
        iData_RowCount = len(pcursrows)
        if iData_RowCount == 0:
            iData_Row = False
        else:
            iData_Row = True
            if iData_Column:
                iData_RowCount += 1
            else:
                iData_ColumnCount = len(pcursrows[0])
        # Define counter min
        iColumn_CountMin = iRange_ColumnCount
        iRow_CountMin = iRange_RowCount
        if iColumn_CountMin > iData_ColumnCount:
            iColumn_CountMin = iData_ColumnCount
        if iRow_CountMin > iData_RowCount:
            iRow_CountMin = iData_RowCount
        elif hshOption["preferstodata"]:
            iSheet_RowCount = sheetTemp.Rows.Count - iRow_Start + 1
            if iSheet_RowCount < iData_RowCount:
                iRow_CountMin = iSheet_RowCount
            else:
                iRow_CountMin = iData_RowCount
        # Define end index
        iRow_End = iRow_Start + (iRow_CountMin - 1)
        iColumn_End = iColumn_Start + (iColumn_CountMin - 1)
        # Check to add column names
        if iData_Column:
            self.CurrentRange_Parts["HEADER"] = sheetTemp.Range(
                sheetTemp.Cells(iRow_Start, iColumn_Start),
                sheetTemp.Cells(iRow_Start, iColumn_End),
            )
            self.CurrentRange_Parts["HEADER"].Value2 = next(zip(*pcurscolumns))
            iRow_Start += 1
            if iRow_Start > iRow_End:
                iData_Row = False
        # Check to add data
        if iData_Row:
            c = int(math.ceil(iRow_CountMin / float(hshOption["pcursorloadmaxrows"])))
            s = e = 0
            for n in range(1, c + 1):
                s, e = e, n * hshOption["pcursorloadmaxrows"]
                iRow_Start_TEMP = iRow_Start + s
                if n == c:
                    iRow_End_TEMP = iRow_End
                else:
                    iRow_End_TEMP = iRow_Start + e - 1
                self.CurrentRange_Parts["DATA"] = sheetTemp.Range(
                    sheetTemp.Cells(iRow_Start_TEMP, iColumn_Start),
                    sheetTemp.Cells(iRow_End_TEMP, iColumn_End),
                )
                values = []
                for row in pcursrows[s:e]:
                    r = []
                    for __, item in enumerate(row):
                        if isinstance(item, datetime.date) or isinstance(item, datetime.datetime):
                            r.append(item.isoformat())
                        else:
                            r.append(item)
                    values.append(r)
                self.CurrentRange_Parts["DATA"].Value2 = values
        else:
            iRow_End = iRow_Start2

        # Enable automatic calculation
        sheetTemp.EnableCalculation = True

        self.CurrentRange_Parts["ALL"] = sheetTemp.Range(
            sheetTemp.Cells(iRow_Start2, iColumn_Start),
            sheetTemp.Cells(iRow_End, iColumn_End),
        )

    # //////////////////////////////////////////////////
    #     TemplateWB_WriteCell
    # //////////////////////////////////////////////////
    def TemplateWB_WriteCell(self, psheet_name, pranges_name, value):
        """
        Writing a cell in excel template

        Args:
            psheet_name (str): the name of target sheet
            pranges_name (str): the name of target ranges
            value (str): data to send
        """
        # Select sheet
        sheetTemp = self.__TemplateWB.Worksheets(psheet_name)
        # Disable automatic calculation : optimization of insertion data
        sheetTemp.EnableCalculation = False
        # copy value
        sheetTemp.Range(pranges_name).Value = value
        # Enable automatic calculation
        sheetTemp.EnableCalculation = True

    # //////////////////////////////////////////////////
    #     TemplateWB_CopyRows
    # //////////////////////////////////////////////////
    def TemplateWB_CopyRows(self, psheet_source, prange_source, psheet_dest, prange_dest):
        """
        Copy Rows in excel template

        Args:
            psheet_source (str): the name of sheet where source is
            prange_source (str): the name of source range
            psheet_dest (str): the name of sheet where to copy
            prange_dest (str): the cell destination
        """
        # Select sheet
        self.log.Debug(f"EXCEL - Copy source {psheet_source} = {prange_source} to {psheet_dest} = {prange_dest}")
        sheetSource = self.__TemplateWB.Worksheets(psheet_source)
        sheetDest = self.__TemplateWB.Worksheets(psheet_dest)
        # Disable automatic calculation : optimization of insertion data
        sheetSource.EnableCalculation = False
        # copy rows
        sheetSource.Rows(prange_source).Copy()
        sheetDest.Activate()
        sheetDest.Range(prange_dest).Select()
        sheetDest.PasteSpecial()
        # Enable automatic calculation
        sheetSource.EnableCalculation = True

    # //////////////////////////////////////////////////
    #     CurrentRange_Borders
    # //////////////////////////////////////////////////
    def CurrentRange_Borders(self, *pborders, ppart="ALL", **kw):
        """
        Excel template save as

        Args:
            ppart (str, optional): the part of range reference. Defaults to "ALL".
            *pborders -- the border identifier list (default='xlEdgeLeft','xlEdgeTop','xlEdgeBottom','xlEdgeRight'
                ,'xlInsideVertical','xlInsideHorizontal')
                Options : LineStyle, ColorIndex, TintAndShade, Weight
        """

        if len(pborders) == 0:
            pborders = [
                "xlEdgeLeft",
                "xlEdgeTop",
                "xlEdgeBottom",
                "xlEdgeRight",
                "xlInsideVertical",
                "xlInsideHorizontal",
            ]

        if isinstance(kw, dict):
            for b in pborders:
                for k, v in kw.items():
                    if isinstance(v, str):
                        v = getattr(self.constants, v)
                    setattr(
                        self.CurrentRange_Parts[ppart].Borders(getattr(self.constants, b)),
                        k,
                        v,
                    )

    # //////////////////////////////////////////////////
    #     TemplateWB_SaveAs
    # //////////////////////////////////////////////////
    def TemplateWB_SaveAs(self, pfile_path):
        """
        Excel template save as

        Args:
            pfile_path (str): the path of new file
        """

        # Managed execution
        try:
            # Optimization parameters reset
            self.__ApplicationXL.ScreenUpdating = True
            self.__ApplicationXL.DisplayStatusBar = True
            self.__ApplicationXL.EnableEvents = True
        # pylint: disable-next=broad-exception-caught
        except Exception as e:
            self.log.Warning("Calculation failed", e, exc_info=1)
        # Managed execution
        try:
            # Refresh data and graphs
            self.__ApplicationXL.CalculateFull()
            self.__TemplateWB.RefreshAll()
        # pylint: disable-next=broad-exception-caught
        except Exception as e:
            # Log
            self.log.Warning("Refresh failed", e, exc_info=1)
        # Save and close
        self.__TemplateWB.SaveAs(pfile_path)
        self.__TemplateWB.Close()

    # //////////////////////////////////////////////////
    #     Close
    # //////////////////////////////////////////////////
    def Close(self):
        """Close Excel application"""
        self.__ApplicationXL.Quit()


# ||||||||||||||||||||||||||||||||||||||||||||||||||
#    Xlsx_read
# ||||||||||||||||||||||||||||||||||||||||||||||||||
class Xlsx_read:
    """Class to read xlsx files."""

    # //////////////////////////////////////////////////
    #     INITIALIZATION
    # //////////////////////////////////////////////////
    def __init__(self):
        # Work variables
        self.__WB = object

        # Start log manager
        self.log = libLog.Log()

    def WB(self):
        """
        return current Workbook

        Returns:
            obj: workbook
        """
        return self.__WB

    # //////////////////////////////////////////////////
    #    Load
    # //////////////////////////////////////////////////
    def Load(self, pWB_path):
        """
        Loading workbook

        Args:
            pWB_path (str): the excel  path

        Raises:
            self.log.CustomException: File does not exist
        """

        # open workbook
        if os.path.isfile(pWB_path):
            with open(pWB_path, "rb") as tmp_file:
                tmp_excel = tmp_file.read()
            # pylint: disable-next=import-outside-toplevel
            from io import BytesIO

            self.__WB = load_workbook(filename=BytesIO(tmp_excel), data_only=True, read_only=True)
        else:
            raise self.log.CustomException(f"File does not exist : {pWB_path}")

    # //////////////////////////////////////////////////
    #     read_data
    # //////////////////////////////////////////////////
    def read_data(self, psheet_name):
        """Return list with content

        Args:
            psheet_name (str or int): name or id of sheet wanted

        Returns:
            tuple:
              - Entete (tuple): list of entetes
              - data (list of tuple) : list of data
        """

        wb_sheet = self.__WB.get_sheet_by_name(psheet_name)
        # on récupère toutes les lignes du documents excel
        datas = []
        for is_first_elt, cells in libTool.signal_first(wb_sheet.values):
            cells = tuple([x for x in cells if x])
            if cells:
                if is_first_elt:
                    Entete = cells
                else:
                    datas.append(cells)
        # on retourne l'objet
        return Entete, datas

    # //////////////////////////////////////////////////
    #     Close
    # //////////////////////////////////////////////////
    def Close(self):
        """Close Excel application"""
        if isinstance(self.__WB, Workbook):
            del self.__WB


# ||||||||||||||||||||||||||||||||||||||||||||||||||
#    Xlsx_write
# ||||||||||||||||||||||||||||||||||||||||||||||||||
class Xlsx_write:
    """Class to write xlsx files."""

    # //////////////////////////////////////////////////
    #     INITIALIZATION
    # //////////////////////////////////////////////////
    def __init__(self):
        # Work variables
        self.__WB = object
        self.CurrentRange_Parts = {"ALL": None, "HEADER": None, "DATA": None}
        self.sheetTemp = None

    # //////////////////////////////////////////////////
    #    TemplateWB_Load
    # //////////////////////////////////////////////////
    def TemplateWB_Load(self, ptemplateWB_path):
        """
        Loading  workbook

        Args:
            ptemplateWB_path (str): the excel template path
        """
        # open workbook
        if os.path.isfile(ptemplateWB_path):
            self.__WB = load_workbook(
                filename=ptemplateWB_path,
                data_only=True,
                read_only=False,
            )
        else:
            self.__WB = Workbook()

    # //////////////////////////////////////////////////
    #     TemplateWB_LoadData
    # //////////////////////////////////////////////////
    def TemplateWB_LoadData(self, psheet_name, pranges_name, pcursrows, pcurscolumns=(), **kw):
        """
        loading data in excel template

        Args:
            psheet_name (str): the name of target sheet
            pranges_name (str): the name of target ranges
            pcursrows (rows): Rows data of cursor
            pcurscolumns (tuple, optional): Columns of cursor. Defaults to ().
            **prowindex -- Start index of row (default=0)
            **pcolindex -- Start index of column (default=0)
            **pdataclean -- Cleaning of old data (default=True)
            **pcursorloadmaxrows -- Cursor to loading : Max rows (default=100000)
            **preferstodata -- New affectation of named range with data size (default=False)
        """
        # ------------
        # init
        # ------------
        hshOption = {
            "prowindex": 0,
            "pcolindex": 0,
            "pdataclean": True,
            "pcursorloadmaxrows": 100000,
            "preferstodata": False,
        }
        self.CurrentRange_Parts["HEADER"] = None
        self.CurrentRange_Parts["DATA"] = None
        self.CurrentRange_Parts["ALL"] = None

        # Setting dictionary option
        if isinstance(kw, dict):
            hshOption.update(kw)

        # Select sheet
        self.sheetTemp = self.__WB[psheet_name]

        # ------------
        # Count col, rows on data
        # ------------
        # ------------
        # Find zone where to write : pranges_name is start, then expand from parameters possibilities
        # ------------
        iData_RowCount = len(pcursrows)
        # Define data counter
        iData_ColumnCount = len(pcurscolumns)
        if iData_ColumnCount == 0:
            iData_Column = False
        else:
            iData_Column = True

        if iData_RowCount == 0:
            iData_Row = False
        else:
            iData_Row = True
            if iData_Column:
                iData_RowCount += 1
            else:
                iData_ColumnCount = len(pcursrows[0])
        # Define Range
        rangeTempInfos = self.__find_range(pranges_name)
        tmp_min_row = rangeTempInfos.min_row + hshOption["prowindex"]
        tmp_max_row = rangeTempInfos.max_row
        if tmp_max_row < (iData_RowCount + tmp_min_row - 1):
            rangeTempInfos.expand(down=iData_RowCount + hshOption["prowindex"] - 1)
            tmp_max_row += iData_RowCount + hshOption["prowindex"] - 1
        rangeTempInfos = rangeTempInfos.intersection(
            CellRange(
                None,
                rangeTempInfos.min_col + hshOption["pcolindex"],
                tmp_min_row,
                rangeTempInfos.max_col,
                tmp_max_row,
            )
        )
        rangeTemp = self.sheetTemp[rangeTempInfos.coord]

        # Delete old data
        if hshOption["pdataclean"]:
            for row in rangeTemp:
                for cell in row:
                    cell.value = None

        # Define Start index
        iColumn_Start = rangeTempInfos.min_col
        iRow_Start = rangeTempInfos.min_row
        iRow_Start2 = rangeTempInfos.min_row
        # Define range counter
        iRange_ColumnCount = rangeTempInfos.size["columns"]
        iRange_RowCount = rangeTempInfos.size["rows"]

        # Define counter min
        iColumn_CountMin = iRange_ColumnCount
        iRow_CountMin = iRange_RowCount
        if iColumn_CountMin > iData_ColumnCount:
            iColumn_CountMin = iData_ColumnCount
        if iRow_CountMin > iData_RowCount:
            iRow_CountMin = iData_RowCount
        elif hshOption["preferstodata"]:
            iSheet_RowCount = self.sheetTemp.max_row - iRow_Start + 1
            if iSheet_RowCount < iData_RowCount:
                iRow_CountMin = iSheet_RowCount
            else:
                iRow_CountMin = iData_RowCount
        # Define end index
        iRow_End = iRow_Start + (iRow_CountMin - 1)
        iColumn_End = iColumn_Start + (iColumn_CountMin - 1)
        # Check to add column names
        styles = []
        if iData_Column:
            self.CurrentRange_Parts["HEADER"] = CellRange(
                None,
                iColumn_Start,
                iRow_Start,
                iColumn_End,
                iRow_End,
            )
            for c in self.CurrentRange_Parts["HEADER"].cells:
                v_row = c[0] - iRow_Start
                v_col = c[1] - iColumn_Start
                self.sheetTemp.cell(row=c[0], column=c[1], value=pcurscolumns[v_col][0])

            iRow_Start += 1
            if iRow_Start > iRow_End:
                iData_Row = False
        # copy style for datas
        for idx in range(iColumn_Start, iColumn_End + 1):
            # pylint: disable-next=protected-access
            styles.append(self.sheetTemp.cell(row=iRow_Start, column=idx)._style)
        for name, r in self.sheetTemp.tables.items():
            rangeTable = CellRange(r)
            if rangeTable.min_col == iColumn_Start and rangeTable.max_col == iColumn_End:
                rangeTable.expand(down=iRow_End - iRow_Start)
                self.sheetTemp.tables[name].ref = rangeTable.coord
                break
        # Check to add data
        if iData_Row:
            c = int(math.ceil(iRow_CountMin / float(hshOption["pcursorloadmaxrows"])))
            s = e = 0
            for n in range(1, c + 1):
                s, e = e, n * hshOption["pcursorloadmaxrows"]
                iRow_Start_TEMP = iRow_Start + s
                if n == c:
                    iRow_End_TEMP = iRow_End
                else:
                    iRow_End_TEMP = iRow_Start + e - 1
                self.CurrentRange_Parts["DATA"] = CellRange(
                    None,
                    iColumn_Start,
                    iRow_Start_TEMP,
                    iColumn_End,
                    iRow_End_TEMP,
                )
                values = []
                for row in pcursrows[s:e]:
                    r = []
                    for __, item in enumerate(row):
                        r.append(item)
                    values.append(r)

                for c in self.CurrentRange_Parts["DATA"].cells:
                    v_row = c[0] - iRow_Start
                    v_col = c[1] - iColumn_Start
                    # pylint: disable-next=protected-access
                    self.sheetTemp.cell(row=c[0], column=c[1])._style = styles[v_col]
                    self.sheetTemp.cell(row=c[0], column=c[1], value=values[v_row][v_col])
                    if isinstance(values[v_row][v_col], datetime.date):
                        self.sheetTemp.cell(row=c[0], column=c[1]).number_format = "dd/mm/yyyy"
                    elif isinstance(values[v_row][v_col], datetime.datetime):
                        self.sheetTemp.cell(row=c[0], column=c[1]).number_format = "dd/mm/yyyy h:mm:ss"
        else:
            iRow_End = iRow_Start2
        self.CurrentRange_Parts["ALL"] = CellRange(
            None,
            iColumn_Start,
            iRow_Start2,
            iColumn_End,
            iRow_End,
        )

    # //////////////////////////////////////////////////
    #     __find_range
    # //////////////////////////////////////////////////
    def __find_range(self, pranges_name):
        try:
            rangeTempInfos = CellRange(pranges_name)
        except ValueError:
            pranges_name = self.__WB.defined_names[pranges_name].attr_text
            if "!" in pranges_name:
                self.sheetTemp = self.__WB[pranges_name.split("!")[0]]
                pranges_name = pranges_name.split("!")[1]
            elif "[]" in pranges_name:
                pranges_name = self.sheetTemp.tables[pranges_name.replace("[]", "")].ref
            rangeTempInfos = self.__find_range(pranges_name)
        return rangeTempInfos

    # //////////////////////////////////////////////////
    #     TemplateWB_WriteCell
    # //////////////////////////////////////////////////
    def TemplateWB_WriteCell(self, psheet_name, pranges_name, value):
        """
        Writing a cell in excel template

        Args:
            psheet_name (str): the name of target sheet
            pranges_name (str): the name of target ranges
            value (str): data to send
        """
        # Select sheet
        sheetTemp = self.__WB[psheet_name]
        # copy value
        sheetTemp[pranges_name] = value

    # //////////////////////////////////////////////////
    #     TemplateWB_CopyRange
    # //////////////////////////////////////////////////
    def TemplateWB_CopyRange(self, psheet_source, prange_source, psheet_dest, prange_dest):
        """
        Copy Range in excel template

        Args:
            psheet_source (str): the name of sheet where source is
            prange_source (str): the name of source range
            psheet_dest (str): the name of sheet where to copy
            prange_dest (str): the cell destination
        """
        # Select sheet
        sheetSource = self.__WB[psheet_source]
        sheetDest = self.__WB[psheet_dest]
        src_rng_infos, dest_rng_infos = CellRange(prange_source), CellRange(prange_dest)
        offset_col = dest_rng_infos.min_col - src_rng_infos.min_col
        offset_row = dest_rng_infos.min_row - src_rng_infos.min_row

        if offset_row > 0:
            for row in rows_from_range(prange_source):
                sheetDest.row_dimensions[coordinate_from_string(row[0])[1] + offset_row] = sheetSource.row_dimensions[
                    coordinate_from_string(row[0])[1]
                ]
        for row in rows_from_range(prange_source):
            for cell in row:
                # pylint: disable-next=protected-access
                sheetDest[cell].offset(row=offset_row, column=offset_col)._style = sheetSource[cell]._style
                sheetDest[cell].offset(row=offset_row, column=offset_col).value = sheetSource[cell].value
        if offset_col > 0:
            for col in cols_from_range(prange_source):
                col_src = column_index_from_string(coordinate_from_string(col[0])[0])
                col_dest = col_src + offset_col
                sheetDest.column_dimensions[get_column_letter(col_dest)].width = sheetSource.column_dimensions[
                    get_column_letter(col_src)
                ].width

    # //////////////////////////////////////////////////
    #     CurrentRange_Borders
    # //////////////////////////////////////////////////
    # TODO Réécrire le code ci dessous avec SC_SUGG_APPRO
    def CurrentRange_Borders(self, *pborders, ppart="ALL", **kw):
        """
        Excel template save as

        Args:
            ppart (str, optional): the part of range reference. Defaults to "ALL".
            *pborders(list): the border identifier list
                (default='xlEdgeLeft','xlEdgeTop','xlEdgeBottom',
                'xlEdgeRight','xlInsideVertical','xlInsideHorizontal')
                Options : 'dashDot','dashDotDot', 'dashed','dotted','double','hair',
                'medium', 'mediumDashDot', 'mediumDashDotDot','mediumDashed',
                'slantDashDot', 'thick', 'thin', ColorIndex, TintAndShade, Weight
        """

        if len(pborders) == 0:
            pborders = [
                "xlEdgeTop",
                "xlEdgeBottom",
                "xlInsideHorizontal",
            ]
        tmp_range = self.sheetTemp[self.CurrentRange_Parts[ppart].coord]

        def applystyle(row, border):
            s = Side(border_style=kw["LineStyle"], color=kw["ColorIndex"])
            for cell in row:
                if border == "TOP/BOTTOM":
                    cell.border = Border(top=s, bottom=s)
                elif "TOP" in border:
                    cell.border = Border(top=s)
                elif "BOTTOM" in border:
                    cell.border = Border(bottom=s)

        if len(tmp_range) > 2 or "xlEdgeTop" not in pborders or "xlEdgeBottom" not in pborders:
            if "xlEdgeTop" in pborders:
                applystyle(tmp_range[0], "TOP")
            if "xlEdgeBottom" in pborders:
                applystyle(tmp_range[-1], "BOTTOM")
        else:
            applystyle(tmp_range[0], "TOP/BOTTOM")
        for row in tmp_range[1:-1]:
            if "xlInsideHorizontal" in pborders:
                applystyle(row, "TOP/BOTTOM")

    # //////////////////////////////////////////////////
    #     row_group
    # //////////////////////////////////////////////////
    def row_group(self, ppart="ALL", hidden=False):
        """
        Excel group rows

        Args:
            ppart (str, optional): the part of range reference. Defaults to "ALL".
            hidden (bool, optional):  Group will be Hidden. Defaults to False.
        """

        self.sheetTemp.row_dimensions.group(
            self.CurrentRange_Parts[ppart].min_row,
            self.CurrentRange_Parts[ppart].max_row,
            outline_level=1,
            hidden=hidden,
        )
        # TODO A écrire
        # https://stackoverflow.com/questions/71257223/how-to-create-new-groups-of-rows-using-openpyxl
        # https://openpyxl.readthedocs.io/en/latest/_modules/openpyxl/worksheet/dimensions.html class DimensionHolder
        # https://openpyxl.readthedocs.io/en/latest/api/openpyxl.worksheet.dimensions.html#openpyxl.worksheet.dimensions.DimensionHolder

    # //////////////////////////////////////////////////
    #     TemplateWB_SaveAs
    # //////////////////////////////////////////////////
    def TemplateWB_SaveAs(self, pfile_path):
        """
        Excel template save as

        Args:
            pfile_path (str): the path of new file
        """
        # Save and close
        self.__WB.template = False
        self.__WB.save(pfile_path)

    # //////////////////////////////////////////////////
    #     Close
    # //////////////////////////////////////////////////
    def Close(self):
        """Close Excel application"""

        if isinstance(self.__WB, Workbook):
            del self.__WB
