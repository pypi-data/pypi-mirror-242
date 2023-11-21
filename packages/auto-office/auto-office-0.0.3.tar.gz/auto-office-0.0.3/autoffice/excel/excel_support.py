'''
-*- coding:utf-8
@Author: GYH
@Software: PyCharm
@Time: 2023/11/7 17:39
@File_name: excel_support.py
'''

import math
import operator
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from .style import *
from ..file.base_support import *


def is_number(cell=None, ws=None, row: int = None, col: int = None):
    '''
    判断一个单元格的值是否为数字, 可识别正常数字、字符串数字

    Args:
        cell: openyxl中的cell单元格对象, 与(ws, row, col)参数组只需导入一个
        ws: openyxl中的worksheet工作表对象, 当cell参数不为空时, 可不导入
        row: int, 行编号, 注意编号不是从0开始的
        col: int, 列编号, 注意编号不是从0开始的
    '''

    cell = ws.cell(row, col) if cell is None else cell

    value = cell.value
    if isinstance(value, (int, float)) or (isinstance(value, str) and str_is_number(value)):
        return True
    else:
        return False


def str2num(cell=None, ws=None, row: int = None, col: int = None):
    '''
    将单元格内以文本形式存在的数字、百分比数字、会计数字转换为数字

    Args:
        cell: openyxl中的cell单元格对象, 与(ws, row, col)参数组只需导入一个
        ws: openyxl中的worksheet工作表对象, 当cell参数不为空时, 可不导入
        row: int, 行编号, 注意编号不是从0开始的
        col: int, 列编号, 注意编号不是从0开始的
    '''

    cell = ws.cell(row, col) if cell is None else cell

    value = cell.value
    if isinstance(value, str) and str_is_number(value):  # 将文本数值转化为数值
        if operator.contains(value, "%"):
            value = float(value.replace("%", "")) / 100  # 去除百分号
            cell.value = value
            cell.number_format = '0.00%'
        elif operator.contains(value, ","):  # 会计数字
            cell.value = float(value.replace(",", ""))
        else:
            cell.value = float(value)


def cell_style(cell, style: dict):
    '''
    将对应单元格对象一次性设置样式
    Args:
        cell: openpyxl下的cell对象
        style: 可在style这个字典中指定好cell各个格式需要的样式。举例:
               style = {
                        'border': 'border_first',
                        'alignment': 'alignment_num',
                        'font': 'font_cn',
                    }
    '''

    for key, value in style.items():
        expr = 'cell.{} = {}'.format(key, value)
        exec(expr)


def adjust_column_dimension(ws, min_col: int, max_col: int):
    '''
    自适应调整从起始列到终点列的的列宽

    Args:
        ws: openyxl中的worksheet工作表对象
        min_col: int, 起始列编号
        max_col: int, 终点列编号
    '''

    column_widths = []       # 记录每一列最终的调整列宽
    for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col)):
        cell_width = []      # 记录每列所有单元格的当前计算长度
        for cell in col:
            if cell.row < MAX_ROWS:  # 最多判断MAX_ROWS行, 防止数据量过大导致崩溃
                value = cell.value
                if value is not None:
                    if isinstance(value, str) is False:
                        value = str(value)

                    char_dic = str_count(value)
                    length = char_dic['cn'] * 5 + char_dic['letter'] / 1.5 \
                             + char_dic['digit'] * 1.2 + char_dic['space'] + char_dic['other'] * 1.1
                    cell_width.append(length)

        cell_width.sort()
        num_extreme = math.ceil(len(cell_width) * 0.01)
        if np.average(cell_width[-num_extreme:]) > 3 * np.average(cell_width[:num_extreme]):  # 当极端值大于三倍均值时, 去除极端值
            cell_width = cell_width[:-num_extreme]
        column_widths.append(max(cell_width) if len(cell_width) > 0 else 8.5)

    for i in range(len(column_widths)):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2  # openpyxl设置字符宽度时会缩水0.5左右个字符，所以+2使左右都空出一个字宽。
        ws.column_dimensions[col_name].width = value if value <= 36 else 36  # 最大列宽上限为36


def df_list2excel(df_list, save_posit, file_name='1', in_one_excel=True):
    '''
    将dataframe构成的列表list导出为excel文件
    Args:
        save_posit: str, excel文件存放文件夹位置
        file_name: 文件名, 当in_one_excel=False时应当提供一个字符串列表, 否则excel文件将分别被命名为1、2、...
        in_one_excel: 是否将dataframe导入一个excel
    '''

    if in_one_excel:
        '''这种情况下要求file_name是一个字符串'''
        path = os.path.join(save_posit, file_name)
        with pd.ExcelWriter(path) as writer:
            for i in range(len(df_list)):
                df_list[i].to_excel(excel_writer=writer, sheet_name='Sheet{}'.format(i+1), index=False)

    if not in_one_excel:
        '''这种情况下要求file_name是一个字符串构成的列表, 若不提供则按照数字命名'''
        if isinstance(file_name, list):
            if len(file_name) == len(df_list):
                for i in range(len(df_list)):
                    path = os.path.join(save_posit, file_name[i])
                    df_list[i].to_excel(path, index=False)
            else:
                raise ValueError("文件名列表长度与文件数量不一致!")
        else:
            for i in range(len(df_list)):
                df_list[i].to_excel(r"{}/{}.xlsx".format(save_posit, i+1), index=False)


def str_count(s):
    '''
    本函数被adjust_column_dimension调用, 用于记录字符串内各类字符的数量
    Args:
        s: str, 输入字符串
    '''
    char_dic = {
        'cn': 0,
        'letter': 0,
        'digit': 0,
        'space': 0,
        'other': 0,
    }
    for char in s:
        if '\u4e00' <= char <= '\u9fa5':  # 判断是否为汉字
            char_dic['cn'] += 1
        elif char.isalpha():  # 判断是否是字母
            char_dic['letter'] += 1
        elif char.isdigit():  # 判断是否是数字
            char_dic['digit'] += 1
        elif char.isspace():  # 判断是否是空格
            char_dic['space'] += 1
        else:
            char_dic['letter'] += 1

    return char_dic