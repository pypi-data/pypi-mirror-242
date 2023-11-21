'''
-*- coding:utf-8
@Author: GYH
@Software: PyCharm
@Time: 2023/11/17 15:31
@File_name: adjust.py
'''


import openpyxl
from .excel_support import *


def excel_format_adj(file, first_line=False):
    '''
    调整excel的格式, 使之变得美观规整
    Args:
        file: str, excel文件的位置路径
        first_line: bool, 当为True时, excel表头将被加入下划线, 默认为false
    '''

    wb = openpyxl.load_workbook(file)

    for ws in wb.worksheets:
        for row, row_ind in zip(ws.iter_rows(), range(1, ws.max_row + 1)):
            for cell in row:
                # 调整以文本形式存在的数字, str2num这个函数就是专门设计出来做这个的
                str2num(ws, cell.row, cell.column)
                style = {
                    'border': 'border_first' if first_line and row_ind == 1 else None,
                    'alignment': 'alignment_num' if is_number(cell) else 'alignment_norm',
                    'font': 'font_cn',
                }
                # 调整样式, cell_style这个函数就是专门设计出来给单元格调整样式的
                cell_style(cell, style)
                # 二次设置字体, 以形成中英文并存的状态
                ws[cell.coordinate].font = font_en
                # 调整行高
                ws.row_dimensions[row_ind].height = 30 if first_line and row_ind == 1 else 16
        # 调整列宽, adjust_column_dimension这个函数设计出来就是专门做这个的
        adjust_column_dimension(ws, 1, ws.max_column)
    wb.save(file)
    wb.close()


def yoy_insert(
    ws = None,
    file: str = None,
    sheet: int = 1,
    rows: list = None,
    yoy_range: list = None,
    mode: str = 'add'
):
    '''
    插入并生成同比增长率行, 但这个方法有一个无法解决的bug, 在"add"模式下, 用openpyxl插入行后, 源文件下方已有的行若存在公式, 不会随着
    程序插入行而自动改变, 这时表格就会出现错误。此时建议由用户提前在要计算yoy的行下手动插入空行, 在程序中使用"fill"模式

    Args:
        ws: worksheet, 当ws为空时代表用文件路径(file参数)获取
        file: 文件路径, 当file为空时应输入ws(worksheet)参数, 二者必有其一
        sheet: 操作的表是第几张sheet, 从1开始计数
        rows: list, 需要计算yoy的数据所在的行编号组成的列表, 举例:[1,3,4]
        yoy_range: 嵌套列表, 例如[[2, 10], [12, 14]]代表要操作的列从第2列开始到第9列, 一级第12列开始到第14列
        mode: 当启用"add"模式时, 程序会自动帮忙创建yoy行, 当启用"fill"模式时, 需要用户自己手动创建行
    '''

    rows.sort()     # 预先进行排序
    if mode == 'add':
       rows = list(map(operator.add, rows, list(range(0, len(rows)))))  # 当为"add"模式时, 依次添加行, 这就需要对row进行累加
    if file:
        wb = openpyxl.load_workbook(file)
        ws = wb.worksheets[sheet - 1]

    for row in rows:
        row_id = row + 1
        if mode == 'add':
            ws.insert_rows(row_id)    # 在目标行的下方补一个yoy行
        ws.cell(row=row_id, column=yoy_range[0][0]-1, value='yoy')   # 在计算范围前写一个标题
        cell_style(
            ws.cell(row=row_id, column=yoy_range[0][0]-1), {'alignment': 'alignment_num', 'font': 'font_blue'}
        )

        for col_range in yoy_range:
            for col_id in range(col_range[0], col_range[1] + 1):
                cell_now = ws.cell(row, col_id)
                cell_pre = ws.cell(row, col_id - 1)
                if isinstance(cell_now.value, (int, float)) and isinstance(cell_pre.value, (int, float)):
                    print('填充所在行:{}, 所在列:{}'.format(row_id, col_id))
                    yoy = "={}/{}-1".format(cell_now.coordinate, cell_pre.coordinate) # 保留计算过程
                    yoy_value = cell_now.value / cell_pre.value - 1
                    ws.cell(row_id, col_id, yoy)
                    ws.cell(row_id, col_id).number_format = '0.00%'
                    ws.cell(row_id, col_id).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                    ws.cell(row_id, col_id).font = Font(name='Calibri', size=10, color='FF0000', italic=True) \
                        if yoy_value < 0 else Font(name='Calibri', size=10, color='000000', italic=True)

    if file:
        wb.save('test.xlsx')
        wb.close()